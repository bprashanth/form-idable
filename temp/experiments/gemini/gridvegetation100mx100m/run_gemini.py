"""
Gemini direct-PDF experiment — no Textract, no crops, one API call.
Usage: python3 run_gemini.py [--model MODEL]
Default model: gemini-3.5-flash
"""

import os, sys, json, time, argparse, re, pathlib
import openpyxl
from openpyxl.styles import Font, PatternFill
from google import genai

# ── args ──────────────────────────────────────────────────────────────────────
parser = argparse.ArgumentParser()
parser.add_argument("--model", default="gemini-3.5-flash")
parser.add_argument("--pdf", default="input.pdf")
parser.add_argument("--out-dir", default=".")
parser.add_argument("--no-thinking", action="store_true",
                    help="Set thinking_budget=0 to disable internal reasoning")
parser.add_argument("--v1-json", default=None,
                    help="Path to Textract v1.json; if given, pass it as extra context")
args = parser.parse_args()

MODEL      = args.model
PDF        = args.pdf
NO_THINK   = args.no_thinking
V1_JSON    = args.v1_json
OUT_DIR    = pathlib.Path(args.out_dir)
OUT_DIR.mkdir(parents=True, exist_ok=True)

suffix = "-nothink" if NO_THINK else ""
suffix += "-v1" if V1_JSON else ""
SLUG = MODEL.replace("/", "-") + suffix

# ── API key ───────────────────────────────────────────────────────────────────
api_key = os.environ.get("GEMINI_API_KEY") or os.environ.get("GOOGLE_API_KEY")
if not api_key:
    env_file = pathlib.Path(
        "/home/desinotorious/src/github.com/bprashanth/good-shepherd/agents/formidable/.env"
    )
    for line in env_file.read_text().splitlines():
        if line.startswith("GEMINI_API_KEY="):
            api_key = line.split("=", 1)[1].strip()
if not api_key:
    sys.exit("GEMINI_API_KEY not found")

client = genai.Client(api_key=api_key)

# ── prompt ────────────────────────────────────────────────────────────────────
PROMPT = """
This is an ecology survey datasheet (PDF). Extract every piece of data from it into
structured JSON. The form may contain multiple repeated form-sheets on the same page.

Return ONLY valid JSON (no markdown fences, no extra text) in this shape:

{
  "sheets": [
    {
      "name": "<section or form name>",
      "rows": [
        ["Header1", "Header2", ...],
        ["value1",  "value2",  ...],
        ...
      ]
    }
  ]
}

Rules:
- One sheet per logical section (header metadata, each table, each free-text block).
- If there are two complete form-sheets on the page (e.g. two grid blocks), emit both
  as separate groups of sheets and prefix each sheet name with the form identifier
  (e.g. "M13: Alien trees", "L13: Alien trees").
- Include printed legends (e.g. cover-scale text) as a single-row sheet named "<form>: Legend".
- Leave cells blank rather than writing "N/A" or "—".
- Do not omit any data, even if it looks like an artifact; flag uncertain reads with a
  trailing " (?)" on the value.
"""

# ── upload & call ─────────────────────────────────────────────────────────────
print(f"[{MODEL}] uploading {PDF}...")
t0 = time.time()

uploaded = client.files.upload(file=PDF, config={"mime_type": "application/pdf"})
print(f"  uploaded: {uploaded.name}")

# build contents list
contents = [uploaded]
if V1_JSON:
    v1_text = pathlib.Path(V1_JSON).read_text()
    contents.append(
        f"\n\nHere is the Textract structured output (v1.json) for this PDF:\n\n"
        f"```json\n{v1_text}\n```\n\n"
        f"Use it as your primary data source, but refer to the PDF itself to "
        f"resolve any ambiguities (especially checkboxes, handwritten marks, "
        f"and margin annotations not captured by Textract)."
    )
contents.append(PROMPT)

# thinking config
gen_config = None
if NO_THINK:
    from google.genai import types as gtypes
    gen_config = gtypes.GenerateContentConfig(
        thinking_config=gtypes.ThinkingConfig(thinking_budget=0)
    )

think_label = "no-thinking" if NO_THINK else "thinking-enabled"
v1_label    = "+v1.json" if V1_JSON else "pdf-only"
print(f"[{MODEL}] calling generate_content ({think_label}, {v1_label})...")
response = client.models.generate_content(
    model=f"models/{MODEL}",
    contents=contents,
    config=gen_config,
)
elapsed = time.time() - t0

# ── token usage & cost ────────────────────────────────────────────────────────
usage = response.usage_metadata
in_tok      = getattr(usage, "prompt_token_count",     0) or 0
out_tok     = getattr(usage, "candidates_token_count", 0) or 0
think_tok   = getattr(usage, "thoughts_token_count",   0) or 0
total       = getattr(usage, "total_token_count", in_tok + out_tok + think_tok) or 0

# per-modality prompt breakdown (IMAGE vs TEXT)
prompt_details = getattr(usage, "prompt_tokens_details", None) or []
image_tok = next((m.token_count for m in prompt_details
                  if hasattr(m, "modality") and "IMAGE" in str(m.modality)), 0)
text_prompt_tok = in_tok - image_tok

# Approximate pricing (USD / 1M tokens).
# Thinking tokens dominate cost for thinking-enabled flash models.
# Calibrated against empirical billing: gemini-3.5-flash run → ₹13.23
# at 84 INR/USD → ~$0.157 with 1881 in + 3901 out + 8795 think tokens.
# → thinking rate implied ~$17.5/M (5× higher than 2.5-flash's $3.50/M).
PRICING = {
    "gemini-3.5-flash":      {"in": 0.15, "out": 0.60, "think": 17.5},
    "gemini-3.1-flash-lite": {"in": 0.10, "out": 0.40, "think":  0.0},
    "gemini-3-flash-preview":{"in": 0.15, "out": 0.60, "think":  3.5},
    "gemini-2.5-flash":      {"in": 0.15, "out": 0.60, "think":  3.5},
    "gemini-2.0-flash-lite": {"in": 0.075,"out": 0.30, "think":  0.0},
}
p = PRICING.get(MODEL, {"in": 0.15, "out": 0.60, "think": 5.0})
cost_usd = (in_tok * p["in"] + out_tok * p["out"] + think_tok * p["think"]) / 1_000_000

meta = {
    "model":             MODEL,
    "pdf":               PDF,
    "elapsed_s":         round(elapsed, 2),
    "prompt_tokens":     in_tok,
    "image_tokens":      image_tok,
    "text_prompt_tokens": text_prompt_tok,
    "output_tokens":     out_tok,
    "thinking_tokens":   think_tok,
    "total_tokens":      total,
    "cost_usd_approx":   round(cost_usd, 6),
    "cost_inr_approx":   round(cost_usd * 84, 4),
    "pricing_note": (f"${p['in']}/M in, ${p['out']}/M out, "
                     f"${p['think']}/M think (empirically calibrated)")
}
print(f"  tokens:  in={in_tok} (image={image_tok} text={text_prompt_tok})  "
      f"out={out_tok}  think={think_tok}  total={total}")
print(f"  cost:    ~${cost_usd:.5f} USD  (~₹{cost_usd*84:.2f} at 84 INR/USD)")
print(f"           breakdown: in=${in_tok*p['in']/1e6:.5f}  "
      f"out=${out_tok*p['out']/1e6:.5f}  think=${think_tok*p['think']/1e6:.5f}")
print(f"  elapsed: {elapsed:.1f}s")

# ── save raw response ─────────────────────────────────────────────────────────
raw_path = OUT_DIR / f"raw_{SLUG}.txt"
raw_path.write_text(response.text)
print(f"  raw response → {raw_path}")

meta_path = OUT_DIR / f"meta_{SLUG}.json"
meta_path.write_text(json.dumps(meta, indent=2))

# ── parse JSON ────────────────────────────────────────────────────────────────
raw = response.text.strip()
# strip accidental markdown fences
raw = re.sub(r"^```(?:json)?\s*", "", raw)
raw = re.sub(r"\s*```$", "", raw)

try:
    data = json.loads(raw)
except json.JSONDecodeError as e:
    print(f"  WARNING: JSON parse failed ({e}). Saving raw only.")
    print("  Trying to extract partial JSON...")
    # try to find the first {...} block
    m = re.search(r'\{.*\}', raw, re.DOTALL)
    if m:
        try:
            data = json.loads(m.group())
        except:
            data = None
    else:
        data = None

if not data or "sheets" not in data:
    print("  Could not extract sheets from response. See raw output.")
    sys.exit(0)

# ── write xlsx ────────────────────────────────────────────────────────────────
xlsx_path = OUT_DIR / f"output_{SLUG}.xlsx"
wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default empty sheet

BOLD  = Font(bold=True)

for sheet_def in data["sheets"]:
    name = re.sub(r'[:\\/?*\[\]]', '-', str(sheet_def.get("name", "Sheet")))[:31]
    rows = sheet_def.get("rows", [])
    ws   = wb.create_sheet(title=name)
    for ri, row in enumerate(rows):
        for ci, val in enumerate(row):
            cell = ws.cell(row=ri + 1, column=ci + 1, value=str(val) if val is not None else "")
            if ri == 0:
                cell.font = BOLD
        # auto-fit column widths (rough)
    for col in ws.columns:
        maxlen = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = min(maxlen + 4, 50)

wb.save(xlsx_path)
print(f"  xlsx → {xlsx_path}  ({len(data['sheets'])} sheets)")

# ── summary ───────────────────────────────────────────────────────────────────
print(f"\n=== Summary for {MODEL} ===")
for s in data["sheets"]:
    rows = s.get("rows", [])
    print(f"  sheet '{s['name']}': {len(rows)} rows")
print(f"\n  Cost: ~${cost_usd:.5f}  |  Tokens: {total}  |  Time: {elapsed:.1f}s")

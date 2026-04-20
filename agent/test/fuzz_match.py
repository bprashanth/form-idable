"""
Step 1: fuzzy-match species values from xlsx against species_name.csv.
Outputs agent/test/fuzz_proposals.json for use in Step 2.

Usage: python3 agent/test/fuzz_match.py
"""
import csv
import json
from pathlib import Path

from openpyxl import load_workbook
from rapidfuzz import fuzz, process

AGENT_ROOT = Path(__file__).resolve().parents[1]
DB_PATH    = AGENT_ROOT / "data" / "species_name.csv"
XLSX_PATH  = AGENT_ROOT / "test" / "output.xlsx"
OUT_PATH   = AGENT_ROOT / "test" / "fuzz_proposals.json"

# ── Load species DB ───────────────────────────────────────────────────────────
# choices: lowercase string -> canonical expanded name
choices = {}
rows = []
with open(DB_PATH, newline="", encoding="utf-8") as f:
    for row in csv.DictReader(f):
        rows.append(row)

for row in rows:
    for field in ["Species name Abbr", "Species name expanded", "Toda name"]:
        val = row[field].strip()
        if val and val != "NA":
            choices[val.lower()] = row["Species name expanded"]

# ── Extract species column values from xlsx ───────────────────────────────────
wb = load_workbook(XLSX_PATH)
ws = wb.active

header_row_idx, header_map = None, {}
for row in ws.iter_rows():
    bold = [(c.column, c.value) for c in row if c.value is not None and c.font and c.font.bold]
    if len(bold) >= 2:
        header_row_idx = row[0].row
        header_map = {col: val for col, val in bold}
        break

spp_col = next((col for col, v in header_map.items()
                if "spp" in str(v).lower() or "name" in str(v).lower()), None)

unique_values = set()
for row in ws.iter_rows(min_row=header_row_idx + 1):
    for c in row:
        if c.column == spp_col and c.value:
            v = str(c.value).strip()
            if v:
                unique_values.add(v)

# ── Fuzzy match ───────────────────────────────────────────────────────────────
proposals = []
for v in sorted(unique_values):
    match = process.extractOne(v.lower(), list(choices.keys()), scorer=fuzz.ratio)
    if match:
        key, score, _ = match
        proposals.append({
            "original": v,
            "candidate": choices[key],
            "matched_key": key,
            "score": round(score, 1),
        })
    else:
        proposals.append({"original": v, "candidate": None, "matched_key": None, "score": 0})

proposals.sort(key=lambda x: -x["score"])

OUT_PATH.write_text(json.dumps({"proposals": proposals}, indent=2), encoding="utf-8")
print(f"Wrote {len(proposals)} proposals to {OUT_PATH}")
for p in proposals:
    print(f"  {p['original']:<22} → {str(p['candidate']):<35} ({p['score']}%)")

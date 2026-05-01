from io import BytesIO
from openpyxl import load_workbook

from config import SYSTEM_SERIAL_HEADER


def _find_header_row(ws):
    """Return (row_idx, {col_int: col_name}) for the first row with 2+ bold cells."""
    for row in ws.iter_rows():
        bold = [(c.column, c.value) for c in row if c.value is not None and c.font and c.font.bold]
        if len(bold) >= 2:
            return row[0].row, {col: val for col, val in bold}
    return None, {}


def get_headers(xlsx_bytes: bytes) -> list:
    wb = load_workbook(BytesIO(xlsx_bytes))
    _, header_map = _find_header_row(wb.active)
    return [str(v) for v in header_map.values() if str(v) != SYSTEM_SERIAL_HEADER]


def extract_species_with_system_serials(xlsx_bytes: bytes, species_cols: list) -> list:
    """
    Walk data rows and collect unique species values with all system_serials
    (from the SYSTEM_SERIAL_HEADER column) where each value appears.

    Returns [{value, system_serials: [int]}] in first-appearance order.
    Raises ValueError if the reserved column is absent.
    """
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    header_row_idx, header_map = _find_header_row(ws)
    if header_row_idx is None:
        return []

    name_to_col = {str(v): k for k, v in header_map.items()}
    species_col_idxs = {name_to_col[n] for n in species_cols if n in name_to_col}
    system_serial_col = name_to_col.get(SYSTEM_SERIAL_HEADER)

    if system_serial_col is None:
        raise ValueError(
            f"Reserved column '{SYSTEM_SERIAL_HEADER}' not found — "
            "workbook must be produced by Good Shepherd upload"
        )

    seen: dict = {}  # value -> [system_serial, ...]

    for row in ws.iter_rows(min_row=header_row_idx + 1):
        row_by_col = {c.column: c for c in row}

        sys_cell = row_by_col.get(system_serial_col)
        if sys_cell is None or sys_cell.value is None:
            continue
        try:
            system_serial = int(sys_cell.value)
        except (ValueError, TypeError):
            continue

        for cell in row:
            if cell.column in species_col_idxs and cell.value is not None:
                v = str(cell.value).strip()
                if v:
                    seen.setdefault(v, []).append(system_serial)

    return [{"value": v, "system_serials": serials} for v, serials in seen.items()]


def apply_species_corrections(xlsx_bytes: bytes, col_names: list, corrections: list) -> bytes:
    """
    corrections: [{original, corrected, system_serials: [int]}]

    Applies each correction only to rows whose SYSTEM_SERIAL_HEADER value is in
    the correction's system_serials list. system_serials is always required.
    """
    # Build lookup: original -> {corrected, system_serials_set}
    correction_map = {}
    for c in corrections:
        if c.get("corrected") and c.get("system_serials"):
            correction_map[c["original"]] = {
                "corrected": c["corrected"],
                "system_serials": set(c["system_serials"]),
            }

    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    header_row_idx, header_map = _find_header_row(ws)
    if header_row_idx is None:
        out = BytesIO()
        wb.save(out)
        return out.getvalue()

    name_to_col = {str(v): k for k, v in header_map.items()}
    target_cols = {name_to_col[n] for n in col_names if n in name_to_col}
    system_serial_col = name_to_col.get(SYSTEM_SERIAL_HEADER)

    for row in ws.iter_rows(min_row=header_row_idx + 1):
        row_by_col = {c.column: c for c in row}

        sys_cell = row_by_col.get(system_serial_col)
        if sys_cell is None or sys_cell.value is None:
            continue
        try:
            system_serial = int(sys_cell.value)
        except (ValueError, TypeError):
            continue

        for cell in row:
            if cell.column in target_cols and cell.value is not None:
                v = str(cell.value).strip()
                if v in correction_map:
                    entry = correction_map[v]
                    if system_serial in entry["system_serials"]:
                        cell.value = entry["corrected"]

    out = BytesIO()
    wb.save(out)
    return out.getvalue()


def apply_serial_numbering(xlsx_bytes: bytes, col_names: list) -> tuple:
    """Returns (corrected_bytes, row_count). Skips the reserved system column."""
    wb = load_workbook(BytesIO(xlsx_bytes))
    ws = wb.active
    header_row_idx, header_map = _find_header_row(ws)
    if header_row_idx is None:
        out = BytesIO()
        wb.save(out)
        return out.getvalue(), 0

    name_to_col = {str(v): k for k, v in header_map.items()}
    target_cols = {
        name_to_col[n] for n in col_names
        if n in name_to_col and n != SYSTEM_SERIAL_HEADER
    }

    count = 0
    for col_idx in target_cols:
        n = 1
        for row in ws.iter_rows(min_row=header_row_idx + 1):
            cell = next((c for c in row if c.column == col_idx), None)
            if cell and cell.value is not None:
                cell.value = n
                n += 1
        count = n - 1

    out = BytesIO()
    wb.save(out)
    return out.getvalue(), count

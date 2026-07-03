import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "v2"

BOLD = Font(bold=True)
TITLE = Font(bold=True, size=12)
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

def setcell(r, c, value, bold=False, flag=False):
    cell = ws.cell(row=r, column=c, value=value)
    if bold:
        cell.font = BOLD
    if flag:
        cell.fill = YELLOW
    return cell

def meta_row(r, label, value, flag=False):
    setcell(r, 1, label, bold=True)
    cell = setcell(r, 2, value)
    if flag:
        for c in range(2, 8):
            ws.cell(row=r, column=c).fill = YELLOW
    return r + 1

def table_header(r, cols):
    for i, h in enumerate(cols):
        setcell(r, 1 + i, h, bold=True)
    return r + 1

def data_row(r, vals, flag_cols=None):
    flag_cols = flag_cols or []
    for i, v in enumerate(vals):
        cell = setcell(r, 1 + i, v)
        if (i + 1) in flag_cols:
            cell.fill = YELLOW
    return r + 1

def quadrant_grid(r, title, nw, ne, sw, se, flag=True):
    setcell(r, 1, title, bold=True, flag=flag)
    r += 1
    setcell(r, 2, "West", bold=True, flag=flag)
    setcell(r, 3, "East", bold=True, flag=flag)
    r += 1
    setcell(r, 1, "North", bold=True, flag=flag)
    setcell(r, 2, nw, flag=flag)
    setcell(r, 3, ne, flag=flag)
    r += 1
    setcell(r, 1, "South", bold=True, flag=flag)
    setcell(r, 2, sw, flag=flag)
    setcell(r, 3, se, flag=flag)
    r += 1
    return r

r = 1

# ========== FORM 1 (Grid M13, top) ==========
setcell(r, 1, "Data sheet for 100 x 100 m grid -- Form 1 (Grid M13)", bold=True)
r += 1

r = meta_row(r, "Grid no:", "M13")
r = meta_row(r, "Date:", "29th Apr (?)", flag=True)
r = meta_row(r, "Observer:", "SK, S, PS (?)", flag=True)
r = meta_row(r, "Grid Centroid GPS:", "10.31305 76.83210 (?)", flag=True)
r = meta_row(r, "Altitude:", "")
r = meta_row(r, "Slope (clinometer):", "15")
r = meta_row(r, "Canopy density (Open<5% / Sparse / Moderate / High):", "Sparse (quadrants below)")
r = meta_row(r, "Canopy composition (Mostly Exotic / Mixed / Mostly Native):", "(quadrants below, NW uncertain)")
r = meta_row(r, "Disturbance:", "Grazing / Firewood / Lopping / Girdling / Tree Cutting / metal wires(?), snares(?)", flag=True)
r += 1

r = quadrant_grid(r, "CAN DENSITY (Grid M13)", "S", "S", "S", "S")
r = quadrant_grid(r, "CAN. COMP (Grid M13)", "2 (?)", "N", "N", "N")
r += 1

# 6) Alien trees (Presence/Absence) -- top
setcell(r, 1, "6) Alien trees (Presence/Absence)", bold=True)
r += 1
r = table_header(r, ["No", "Species", "Quarter1", "Quarter2", "Quarter3", "Quarter4", "Comments / notes"])
r = data_row(r, ["1", "Silver oak", "[X]", "[X]", "[X]", "Y [X]", "[ ]"], flag_cols=[6])
r = data_row(r, ["2", "Maesopsis", "[X]", "[X]", "[X]", "[X]", "[ ]"])
r = data_row(r, ["3", "Spathodea", "[X]", "[X]", "[X]", "[X]", "[ ]"])
r = data_row(r, ["4", "Eucalyptus", "[X]", "[X]", "[X]", "[X]", "[ ]"])
r = data_row(r, ["5", "Other*", "", "[ ]", "[ ]", "[ ]", "[ ]"])
r += 1
r = meta_row(r, "* Other species:", "Coffee")
r += 1

# 7) Alien plant prevalence -- top
setcell(r, 1, "7) Alien plant prevalence", bold=True)
r += 1
setcell(r, 1, "Cover: 0=Absent, 1=Low <33%, 2=Med 33-67%, 3=High >67%; Quarters: note P/A")
r += 1
r = table_header(r, ["Species", "Cover", "Quarter1", "Quarter2", "Quarter3", "Quarter4", "Comments / notes"])
r = data_row(r, ["Lantana", "", "X", "X", "y [X]", "[X]", ""], flag_cols=[5])
r = data_row(r, ["Chromolaena", "low", "[X]", "[X]", "", "[X]", ""])
r = data_row(r, ["Mikania", "lm", "[X]", "[X]", "[X]", "", ""])
r = data_row(r, ["Wedelia", "", "[X]", "Y", "[X]", "X", ""], flag_cols=[4, 6])
r = data_row(r, ["Montanoa", "", "7", "X", "x", "", ""], flag_cols=[3, 4, 5])
r = data_row(r, ["Gliricidia", "100", "[X]", "[X]", "X", "[X]", ""])
r = data_row(r, ["Pohyjorum", "lon", "[X]", "[X]", "[X]", "[X]", ""])
r = data_row(r, ["", "", "", "", "", "", ""])
r = data_row(r, ["", "", "", "", "", "", ""])
r = data_row(r, ["Other*", "", "", "", "[ ]", "[ ]", ""])
r += 1
r = meta_row(r, "* Other species:", "")
r += 2

# ========== FORM 2 (Grid L13, bottom) ==========
setcell(r, 1, "Data sheet for 100 x 100 m grid -- Form 2 (Grid L13)", bold=True)
r += 1

r = meta_row(r, "Grid no:", "L13")
r = meta_row(r, "Date:", "29th Apr (?)", flag=True)
r = meta_row(r, "Observer:", "SK, S, PS")
r = meta_row(r, "Grid Centroid GPS:", "10.31309 76.83302")
r = meta_row(r, "Altitude:", "")
r = meta_row(r, "Slope (clinometer):", "20 (?)", flag=True)
r = meta_row(r, "Canopy density (Open<5% / Sparse / Moderate / High):", "Sparse (?) (quadrants below)", flag=True)
r = meta_row(r, "Canopy composition (Mostly Exotic / Mixed / Mostly Native):", "Mixed (?) (quadrants below)", flag=True)
r = meta_row(r, "Disturbance:", "(not captured -- possible Textract gap, check crop)", flag=True)
r += 1

r = quadrant_grid(r, "CAN DENSITY (Grid L13)", "S (?)", "S (?)", "S (?)", "S (?)")
r = quadrant_grid(r, "CAN. COMP (Grid L13)", "M (?)", "M (?)", "M (?)", "M (?)")
r += 1

# 6) Alien trees (Presence/Absence) -- bottom
setcell(r, 1, "6) Alien trees (Presence/Absence)", bold=True)
r += 1
r = table_header(r, ["No", "Species", "Quarter1", "Quarter2", "Quarter3", "Quarter4", "Comments / notes"])
r = data_row(r, ["1", "Silver oak", "X", "[X]", "[X]", "[X]", "[ ]"])
r = data_row(r, ["2", "Maesopsis", "", "[X]", "[X]", "[X]", "[ ]"])
r = data_row(r, ["3", "Spathodea", "&", "X", "[X]", "[X]", "[ ]"], flag_cols=[3])
r = data_row(r, ["4", "Eucalyptus", "X", "X", "X", "[X] r", "[ ]"], flag_cols=[6])
r = data_row(r, ["5", "Other*", "", "[ ]", "[ ]", "[ ]", "[ ]"])
r += 1
r = meta_row(r, "* Other species:", "coffee.")
r += 1

# 7) Alien plant prevalence -- bottom
setcell(r, 1, "7) Alien plant prevalence", bold=True)
r += 1
setcell(r, 1, "Cover: 0=Absent, 1=Low <33%, 2=Med 33-67%, 3=High >67%; Quarters: note P/A")
r += 1
r = table_header(r, ["Species", "Cover", "Quarter1", "Quarter2", "Quarter3", "Quarter4", "Comments / notes"])
r = data_row(r, ["Lantona", "low", "[X]", "", "[X] A", "[X]", "[ ]"], flag_cols=[5])
r = data_row(r, ["Chromolaena", "low", "[X]", "[X]", "[X]", "[X]", "[ ]"])
r = data_row(r, ["Mikania", "low", "[X]", "[X]", "[X]", "[X]", "[ ]"])
r = data_row(r, ["Wedelia", "Httigh (?)", "[X]", "[X]", "", "[X]", ""], flag_cols=[2])
r = data_row(r, ["Montanaa", "", "[X]", "[X]", "X", "X", "[ ]"], flag_cols=[5, 6])
r = data_row(r, ["Gliricidia", "low", "[X]", "[X]", "[X]", "[X]", ""])
r = data_row(r, ["Polygonum", "how (?)", "[X]", "[X]", "[X]", "", "[ ]"], flag_cols=[2])
r = data_row(r, ["", "", "[ ]", "[ ]", "[ ]", "[ ]", "[ ]"])
r = data_row(r, ["", "", "", "", "[ ]", "[ ]", ""])
r = data_row(r, ["Other*", "", "[ ]", "[ ]", "[ ]", "[ ]", "[ ]"])
r += 1
r = meta_row(r, "* Other species:", "")

# column widths
for col, w in zip("ABCDEFG", [42, 28, 12, 12, 12, 12, 18]):
    ws.column_dimensions[col].width = w

wb.save("golden.xlsx")
print("done, rows =", r)

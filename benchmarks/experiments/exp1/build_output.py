import json
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)

# ── colour helpers ──────────────────────────────────────────────────────────
YELLOW = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
LIGHT_BLUE = PatternFill(start_color="DDEEFF", end_color="DDEEFF", fill_type="solid")
GREY = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
BOLD = Font(bold=True)
BOLD_LARGE = Font(bold=True, size=12)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

thin = Side(style="thin")
THIN_BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)

def set_header(ws, row, col, text, fill=GREY):
    c = ws.cell(row=row, column=col, value=text)
    c.font = BOLD
    c.fill = fill
    c.alignment = LEFT
    c.border = THIN_BORDER
    return c

def set_data(ws, row, col, text, flag=False):
    c = ws.cell(row=row, column=col, value=text)
    c.alignment = LEFT
    c.border = THIN_BORDER
    if flag:
        c.fill = YELLOW
    return c

def set_label(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = BOLD
    c.alignment = LEFT
    return c

def set_value(ws, row, col, text, flag=False):
    c = ws.cell(row=row, column=col, value=text)
    c.alignment = LEFT
    if flag:
        c.fill = YELLOW
    return c

# ── load v1.json ────────────────────────────────────────────────────────────
with open("v1.json") as f:
    v1 = json.load(f)

tables    = v1["tables"]
kv        = v1["key_values"]
other     = v1["other_text"]

# convenience: key_values as dict
kv_dict = {item["key"]: item for item in kv}

# ── create workbook ─────────────────────────────────────────────────────────
wb = openpyxl.Workbook()
ws = wb.active
ws.title = "v2"

# column widths
ws.column_dimensions["A"].width = 22
ws.column_dimensions["B"].width = 30
ws.column_dimensions["C"].width = 16
ws.column_dimensions["D"].width = 16
ws.column_dimensions["E"].width = 16
ws.column_dimensions["F"].width = 16
ws.column_dimensions["G"].width = 16
ws.column_dimensions["H"].width = 22

# ═══════════════════════════════════════════════════════════════════════════
# ROW 1-2 : PROJECT TITLE
# ═══════════════════════════════════════════════════════════════════════════
r = 1
c1 = ws.cell(row=r, column=1,
             value="Ecological recovery in restored areas project")
c1.font = BOLD_LARGE
c1.alignment = CENTER
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

r = 2
c2 = ws.cell(row=r, column=1,
             value="20 X 20 m vegetation plot data sheet")
c2.font = Font(bold=True, size=11)
c2.alignment = CENTER
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

ws.row_dimensions[1].height = 20
ws.row_dimensions[2].height = 18

# ═══════════════════════════════════════════════════════════════════════════
# ROWS 4-9 : METADATA  (from key_values + other_text)
# ═══════════════════════════════════════════════════════════════════════════
r = 4

# Date | Site ID | Plot ID
set_label(ws, r, 1, "Date")
date_val = kv_dict.get("Date", {}).get("value", "")
set_value(ws, r, 2, date_val)

set_label(ws, r, 3, "Site ID")
site_val = kv_dict.get("Site ID", {}).get("value", "")
set_value(ws, r, 4, site_val)

set_label(ws, r, 5, "Plot ID")
plot_val = kv_dict.get("Plot ID", {}).get("value", "")
set_value(ws, r, 6, plot_val)

r = 5
# Centroid long. X / lat. Y – Textract read the value as one string in other_text
# The combined coordinate "10.30217,76.84301" spans both fields.
coords = next((x["text"] for x in other if "10.30217" in x["text"]), "")
set_label(ws, r, 1, "Centroid long. X")
set_value(ws, r, 2, coords, flag=(coords == ""))   # flag if missing

set_label(ws, r, 3, "Centroid lat. Y")
set_value(ws, r, 4, coords, flag=(coords == ""))

r = 6
set_label(ws, r, 1, "Data collectors")
dc_val = kv_dict.get("Data collectors", {}).get("value", "")
set_value(ws, r, 2, dc_val)

r = 7
set_label(ws, r, 1, "Canopy densitometer reading")
# no value captured by Textract – flag for reviewer
set_value(ws, r, 2, "", flag=True)

r = 8
set_label(ws, r, 1, "Canopy (Open / Closed)")
# key_value: "Closed)" -> "[X]"  means Closed is checked
closed_val = kv_dict.get("Closed)", {}).get("value", "")
set_value(ws, r, 2, closed_val if closed_val else "Closed [X]")

r = 9
set_label(ws, r, 1, "Page of")
pof_val = kv_dict.get("Page of", {}).get("value", "")
set_value(ws, r, 2, pof_val, flag=True)

# ═══════════════════════════════════════════════════════════════════════════
# ROWS 11-16 : GROUND COVER GRID  (new section – from other_text)
# ═══════════════════════════════════════════════════════════════════════════
# Textract did NOT recognise this as a table; recovered from other_text layout.
# Column labels (y ≈ 0.31-0.32): Rock Soil, Litter, Grass, Wedelia, Herbs, Copre(?)
# Row labels   (y ≈ 0.33-0.40): Absent, Low <33%, Med 33-67%, High >67%
# Data cells LEFT BLANK and YELLOW-flagged for human reviewer.

r = 11
gc_title = ws.cell(row=r, column=1, value="Ground cover")
gc_title.font = BOLD_LARGE
gc_title.alignment = LEFT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=8)

gc_col_labels = ["Rock Soil", "Litter", "Grass", "Wedelia", "Herbs", "Copre(?)"]
gc_row_labels = ["Absent", "Low <33%", "Med 33-67%", "High >67%"]

r = 12
# header row
set_header(ws, r, 1, "Cover class")
for ci, lbl in enumerate(gc_col_labels, start=2):
    set_header(ws, r, ci, lbl)

for ri, row_lbl in enumerate(gc_row_labels, start=1):
    r = 12 + ri
    set_header(ws, r, 1, row_lbl)
    for ci in range(2, 2 + len(gc_col_labels)):
        set_data(ws, r, ci, "", flag=True)   # yellow – reviewer fills from scan

gc_section_start = 11
gc_section_end   = 12 + len(gc_row_labels)   # = 16

# blank row
ws.cell(row=17, column=1, value="")

# ═══════════════════════════════════════════════════════════════════════════
# ROW 18 : TREES heading
# ═══════════════════════════════════════════════════════════════════════════
r = 18
trees_title = ws.cell(row=r, column=1, value="Trees")
trees_title.font = BOLD_LARGE
trees_title.alignment = LEFT
ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=5)

# ═══════════════════════════════════════════════════════════════════════════
# ROW 19 : TREES table header row
# ═══════════════════════════════════════════════════════════════════════════
tree_headers = ["No.", "Species", "GBH (cm)", "Height (m)", "Remarks"]
r = 19
for ci, h in enumerate(tree_headers, start=1):
    set_header(ws, r, ci, h)

# ═══════════════════════════════════════════════════════════════════════════
# ROWS 20+ : TREES data  (from v1.tables[0])
# ═══════════════════════════════════════════════════════════════════════════
tree_table = tables[0]
cells_by_row = {}
for cell in tree_table["cells"]:
    cells_by_row.setdefault(cell["r"], {})[cell["c"]] = cell

# rows 2-27 are data rows (row 1 = header)
tree_data_start = 20
for trow in range(2, tree_table["n_rows"] + 1):
    row_cells = cells_by_row.get(trow, {})
    out_row = tree_data_start + (trow - 2)
    for tcol in range(1, tree_table["n_cols"] + 1):
        cell = row_cells.get(tcol)
        if cell:
            txt = cell["text"]
            conf = cell.get("conf", 100)
            flag = conf < 75 or txt in ('"', "'", "\"", "11")
            # ditto marks are valid transcription – keep but flag for reviewer
            set_data(ws, out_row, tcol, txt, flag=flag)
        else:
            set_data(ws, out_row, tcol, "", flag=True)

# ═══════════════════════════════════════════════════════════════════════════
# Freeze panes at header row of trees table
# ═══════════════════════════════════════════════════════════════════════════
ws.freeze_panes = "A20"

# ── save ────────────────────────────────────────────────────────────────────
wb.save("output.xlsx")
print("output.xlsx written.")

# ── v2_meta.json – record new sections added beyond v1 ─────────────────────
import json as _json

# Ground-cover grid was entirely recovered from other_text (not in v1 tables/kv)
# Bounding boxes of the constituent other_text entries:
# Rock Soil [0.237,0.313], Litter [0.365,0.316], Grass [0.438,0.317],
# Wedelia [0.509,0.318], Herbs [0.59,0.319], Copre [0.663,0.324]
# Absent [0.11,0.332], Low [0.11,0.353], Med 33-67% [0.11,0.375],
# High [0.109,0.396]  =>  bbox spans x:[0.11,0.716], y:[0.313,0.410]
meta = {
    "new_sections": [
        {
            "sheet": "v2",
            "rows": [gc_section_start, gc_section_end],
            "bbox": [0.11, 0.313, 0.716, 0.410]
        }
    ]
}
with open("v2_meta.json", "w") as f:
    _json.dump(meta, f, indent=2)
print("v2_meta.json written.")

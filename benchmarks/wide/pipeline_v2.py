#!/usr/bin/env python3
"""Local, opt-in Formidable v2 orchestrator. No AWS operations.

One generic structure pass is used both to confirm template identity and, after
an abstention, as the schema for generic extraction. Known templates receive an
exact lattice; unknown or mixed documents safely remain canonical.
"""
from __future__ import annotations

import argparse
import copy
import json
from pathlib import Path

import fitz
import openpyxl

import canonical
import ecology_review
import review_manifest
import structured_pipeline
import template_labels
import template_match
import template_pipeline


def registries(manifest_path: Path, templates: Path):
    manifest = json.loads(manifest_path.read_text())
    pixels, labels = {}, {}
    for item in manifest:
        descriptor = template_match.TemplatePage(item["template"], int(item["page"]) - 1)
        pixels[descriptor] = template_match.render(
            templates / descriptor.template, descriptor.page)
        labels[descriptor.key] = template_labels.template_tokens(
            templates / descriptor.template, descriptor.page)
    return manifest, pixels, labels


def route_document(form_dir: Path, structure_tag: str, schema_model: str,
                   templates: Path, manifest_path: Path,
                   *, reuse_structure=False):
    manifest, pixel_registry, label_registry = registries(manifest_path, templates)
    page_count = fitz.open(form_dir / "input.pdf").page_count
    structure_dir = form_dir / "canonical_outputs" / structure_tag
    required = [structure_dir / f"page_{page}__structure__{schema_model}.json"
                for page in range(1, page_count + 1)]
    if not reuse_structure or not all(path.exists() for path in required):
        structured_pipeline.prepare_structures(
            form_dir, schema_model, structure_tag, reuse=reuse_structure)
    pages = []
    for page_number, structure_path in enumerate(required, 1):
        ranking = template_match.rank(
            template_match.render(form_dir / "input.pdf", page_number - 1), pixel_registry)
        pixel = template_match.verdict(ranking)
        label = None
        if pixel["status"] == "known_template":
            raw = json.loads(structure_path.read_text())
            label = template_labels.confirm(pixel["candidate"], raw, label_registry)
        accepted = pixel["status"] == "known_template" and label and label["accepted"]
        pages.append({
            "input_page": page_number,
            "status": "known_template" if accepted else "unknown_template",
            "template": pixel.get("template") if accepted else None,
            "template_page": pixel.get("page") if accepted else None,
            "pixel": pixel, "printed_labels": label,
        })
    all_known = all(page["status"] == "known_template" for page in pages)
    route = {
        "version": "formidable-route-v1", "source": str(form_dir / "input.pdf"),
        "policy": "exact route only when every page clears pixel and printed-label gates",
        "document_route": "known_template" if all_known else "generic_canonical",
        "mixed_page_policy": "generic_canonical",
        "registry": str(manifest_path), "registry_pages": len(manifest), "pages": pages,
    }
    return route


def merge_workbooks(paths, destination):
    output = openpyxl.Workbook()
    output.remove(output.active)
    for page_number, path in enumerate(paths, 1):
        source = openpyxl.load_workbook(path)
        old = source.active
        new = output.create_sheet(f"page{page_number}")
        new.sheet_view.showGridLines = old.sheet_view.showGridLines
        new.page_setup = copy.copy(old.page_setup)
        new.page_margins = copy.copy(old.page_margins)
        new.print_options = copy.copy(old.print_options)
        new.sheet_properties = copy.copy(old.sheet_properties)
        for row in old.iter_rows():
            for cell in row:
                target = new[cell.coordinate]
                target.value = cell.value
                if cell.has_style:
                    # Style IDs are workbook-local. Copy public components so
                    # the destination registers its own font/fill/border IDs.
                    target.font = copy.copy(cell.font)
                    target.fill = copy.copy(cell.fill)
                    target.border = copy.copy(cell.border)
                    target.alignment = copy.copy(cell.alignment)
                    target.protection = copy.copy(cell.protection)
                    target.number_format = cell.number_format
                if cell.comment:
                    target.comment = copy.copy(cell.comment)
                if cell.hyperlink:
                    target._hyperlink = copy.copy(cell.hyperlink)
        for merged in old.merged_cells.ranges:
            new.merge_cells(str(merged))
        for key, dimension in old.column_dimensions.items():
            new.column_dimensions[key].width = dimension.width
        for key, dimension in old.row_dimensions.items():
            new.row_dimensions[key].height = dimension.height
    output.save(destination)


def combine_template_manifests(manifests, route):
    cells, attention, readers = [], [], []
    for page_number, manifest in enumerate(manifests, 1):
        readers.append({"page": page_number, **manifest["readers"]})
        id_map = {}
        for cell in manifest["cells"]:
            old = cell["id"]
            cell["id"] = f"p{page_number}:{old}"
            cell["page"] = page_number
            id_map[old] = cell["id"]
            cells.append(cell)
        for item in manifest["views"]["transcription_attention"]:
            item["cell_id"] = id_map[item["cell_id"]]
            item["page"] = page_number
            attention.append(item)
    return {
        "version": review_manifest.VERSION, "route": route,
        "policy": {
            "literal_transcription_is_immutable": True,
            "peer_readers_select_review_regions_not_replacements": True,
            "ecology_suggestions_are_separate": True,
        },
        "readers_by_page": readers,
        "summary": {"target_cells_including_blanks": len(cells),
                    "reader_disagreements": sum(
                        cell["status"] == "reader_disagreement" for cell in cells),
                    "low_confidence_primary": sum(
                        cell["status"] == "low_confidence" for cell in cells),
                    "transcription_review_cells": len(attention),
                    "ecology_findings": 0},
        "cells": cells,
        "views": {"transcription_attention": attention, "ecology_anomalies": []},
    }


def ecology_for_template(manifest, cache, online=False):
    records = [ecology_review.Record(
        {"page": cell["page"], "cell": cell["id"]},
        cell.get("context") or "unlabelled", cell.get("presented_value"))
        for cell in manifest["cells"] if cell.get("presented_value") not in (None, "")]
    findings = ecology_review.numeric_findings(records)
    if online:
        findings += ecology_review.taxonomy_findings(
            records, ecology_review.GBIFClient(cache))
    return findings


def run_known(form_dir, output, route, templates, primary, peer, ecology_online):
    workbook_paths, manifests, reports = [], [], []
    for page in route["pages"]:
        page_dir = output / "template_pages" / f"page_{page['input_page']:03d}"
        primary_dir, peer_dir = page_dir / "primary", page_dir / "peer"
        template = templates / page["template"]
        primary_report = template_pipeline.run(
            form_dir, template, page["template_page"] - 1, primary, "page", "primary",
            input_page=page["input_page"] - 1, output_dir=primary_dir)
        peer_report = None
        if peer:
            peer_report = template_pipeline.run(
                form_dir, template, page["template_page"] - 1, peer, "page", "peer",
                input_page=page["input_page"] - 1, output_dir=peer_dir)
        page_manifest = review_manifest.from_template(
            primary_dir, peer_dir if peer else None, route=page)
        manifests.append(page_manifest)
        workbook_paths.append(primary_dir / "output.xlsx")
        reports.append({"page": page["input_page"], "primary": primary_report,
                        "peer": peer_report})
    merge_workbooks(workbook_paths, output / "output.xlsx")
    manifest = combine_template_manifests(manifests, route)
    findings = ecology_for_template(
        manifest, output / ".cache" / "gbif", online=ecology_online)
    manifest["views"]["ecology_anomalies"] = [
        {"finding_id": index + 1, **finding} for index, finding in enumerate(findings)]
    manifest["summary"]["ecology_findings"] = len(findings)
    (output / "ecology_review.json").write_text(json.dumps({
        "version": "formidable-ecology-review-v1",
        "policy": "flags and suggestions only; literal values are never changed",
        "findings": findings}, indent=2, ensure_ascii=False) + "\n")
    (output / "review_manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n")
    return {"route": "known_template", "pages": reports,
            "review": manifest["summary"]}


def run_generic(form_dir, output, route, structure_tag, schema_model,
                models, ecology_online):
    report = structured_pipeline.run(
        form_dir, schema_model, models, structure_tag, reuse_structure=True)
    canonical_dir = form_dir / "canonical_outputs" / structure_tag
    document = json.loads((canonical_dir / "canonical.json").read_text())
    records = ecology_review.canonical_records(document)
    findings = ecology_review.numeric_findings(records)
    if ecology_online:
        findings += ecology_review.taxonomy_findings(
            records, ecology_review.GBIFClient(output / ".cache" / "gbif"))
    ecology_report = {
        "version": "formidable-ecology-review-v1",
        "policy": "flags and suggestions only; literal values are never changed",
        "findings": findings,
    }
    ecology_review.apply_findings(document, findings)
    canonical.dump(document, output / "canonical.json")
    canonical.write_xlsx(document, output / "output.xlsx")
    ecology_review.add_review_sheet(output / "output.xlsx", findings)
    (output / "ecology_review.json").write_text(
        json.dumps(ecology_report, indent=2, ensure_ascii=False) + "\n")
    manifest = review_manifest.from_canonical(document, ecology_report)
    manifest["route"] = route
    (output / "review_manifest.json").write_text(
        json.dumps(manifest, indent=2, ensure_ascii=False) + "\n")
    return {"route": "generic_canonical", "extraction": report,
            "review": manifest["summary"]}


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("form", type=Path,
                        help="directory containing input.pdf")
    parser.add_argument("--templates", type=Path, required=True)
    parser.add_argument("--registry", type=Path, required=True)
    parser.add_argument("--tag", default="formidable_v2")
    parser.add_argument("--schema-model", default="gemini-3.5-flash")
    parser.add_argument("--primary-model", default="gemini-3.6-flash")
    parser.add_argument("--peer-model", default="gemini-3.5-flash",
                        help="empty string disables peer reader")
    parser.add_argument("--reuse-structure", action="store_true",
                        help="route from already saved structure JSON; make no structure call")
    parser.add_argument("--route-only", action="store_true")
    parser.add_argument("--ecology-online", action="store_true",
                        help="enable cached GBIF taxonomy lookups")
    args = parser.parse_args()
    form_dir = args.form.resolve()
    output = form_dir / "v2_outputs" / args.tag
    output.mkdir(parents=True, exist_ok=True)
    route = route_document(form_dir, args.tag, args.schema_model,
                           args.templates.resolve(), args.registry.resolve(),
                           reuse_structure=args.reuse_structure)
    (output / "route.json").write_text(json.dumps(route, indent=2) + "\n")
    if args.route_only:
        print(json.dumps(route, indent=2))
        return
    if route["document_route"] == "known_template":
        report = run_known(form_dir, output, route, args.templates.resolve(),
                           args.primary_model, args.peer_model.strip(), args.ecology_online)
    else:
        report = run_generic(
            form_dir, output, route, args.tag, args.schema_model,
            [args.peer_model, args.primary_model] if args.peer_model.strip()
            else [args.primary_model], args.ecology_online)
    report["output"] = str(output)
    (output / "run.json").write_text(json.dumps(report, indent=2) + "\n")
    print(json.dumps(report, indent=2))


if __name__ == "__main__":
    main()

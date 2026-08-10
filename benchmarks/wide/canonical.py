#!/usr/bin/env python3
"""Canonical, source-linked representation for form extraction experiments."""
from __future__ import annotations

import json
import math
import re
from collections import Counter
from pathlib import Path
from typing import Any

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


VERSION = "formidable-canonical-v1"
GREEN = PatternFill("solid", fgColor="C6EFCE")
YELLOW = PatternFill("solid", fgColor="FFF2CC")
RED = PatternFill("solid", fgColor="F4CCCC")
ORANGE = PatternFill("solid", fgColor="FCE4D6")


def slug(text: str, fallback: str) -> str:
    value = re.sub(r"[^a-z0-9]+", "_", str(text).casefold()).strip("_")
    return value or fallback


def bbox(value) -> list[float] | None:
    if not isinstance(value, list) or len(value) != 4:
        return None
    try:
        coords = [float(x) for x in value]
    except (TypeError, ValueError):
        return None
    if max(coords) > 1:
        scale = max(coords)
        coords = [x / scale for x in coords]
    x0, y0, x1, y1 = coords
    x0, x1 = sorted((max(0.0, min(1.0, x0)), max(0.0, min(1.0, x1))))
    y0, y1 = sorted((max(0.0, min(1.0, y0)), max(0.0, min(1.0, y1))))
    if x1 - x0 < 0.001 or y1 - y0 < 0.001:
        return None
    return [round(x0, 6), round(y0, 6), round(x1, 6), round(y1, 6)]


def row_bbox(value) -> list[float] | None:
    """Repair common coordinate-order drift using row shape, not content."""
    if not isinstance(value, list) or len(value) != 4:
        return None
    a, b, c, d = value
    candidates = ([a, b, c, d], [b, a, d, c],
                  [b, a, c, d], [a, b, d, c])
    regions = [region for region in (bbox(candidate) for candidate in candidates)
               if region is not None]
    if not regions:
        return None
    # Physical table rows are wide and short. This also handles providers that
    # inconsistently emit y/x order between pages.
    return max(regions, key=lambda region:
               (region[2] - region[0]) / max(1e-6, region[3] - region[1]))


def normalize_structure(raw: dict[str, Any], page_number: int) -> dict[str, Any]:
    """Repair harmless model formatting drift while rejecting ambiguous IDs."""
    page = {
        "page_number": page_number,
        "metadata_fields": [],
        "tables": [],
        "free_text_regions": [],
    }
    used_fields = set()
    for index, field in enumerate(raw.get("metadata_fields") or []):
        label = str(field.get("label") or "").strip()
        region = bbox(field.get("bbox"))
        if not label or region is None:
            continue
        field_id = slug(field.get("id") or label, f"field_{index + 1}")
        while field_id in used_fields:
            field_id += "_2"
        used_fields.add(field_id)
        page["metadata_fields"].append({"id": field_id, "label": label, "bbox": region})

    used_tables = set()
    for table_index, table in enumerate(raw.get("tables") or []):
        region = bbox(table.get("bbox"))
        columns = []
        used_columns = set()
        for col_index, column in enumerate(table.get("columns") or []):
            label = str(column.get("label") or "").strip() or f"Column {col_index + 1}"
            column_id = slug(column.get("id") or label, f"column_{col_index + 1}")
            while column_id in used_columns:
                column_id += "_2"
            used_columns.add(column_id)
            try:
                x0, x1 = float(column.get("x0")), float(column.get("x1"))
            except (TypeError, ValueError):
                x0 = col_index / max(1, len(table.get("columns") or []))
                x1 = (col_index + 1) / max(1, len(table.get("columns") or []))
            if max(x0, x1) > 1:
                scale = max(x0, x1)
                x0, x1 = x0 / scale, x1 / scale
            x0, x1 = sorted((max(0.0, min(1.0, x0)), max(0.0, min(1.0, x1))))
            columns.append({
                "id": column_id,
                "label": label,
                "parent": str(column.get("parent") or "").strip() or None,
                "value_kind": str(column.get("value_kind") or "unknown").strip().casefold(),
                "x0": round(x0, 6),
                "x1": round(x1, 6),
            })
        if region is None or not columns:
            continue
        table_id = slug(table.get("id") or table.get("title"), f"table_{table_index + 1}")
        while table_id in used_tables:
            table_id += "_2"
        used_tables.add(table_id)
        page["tables"].append({
            "id": table_id,
            "title": str(table.get("title") or "").strip(),
            "bbox": region,
            "estimated_rows": max(0, int(table.get("estimated_rows") or 0)),
            "columns": columns,
            "rows": [],
        })

    for index, item in enumerate(raw.get("free_text_regions") or []):
        region = bbox(item.get("bbox"))
        if region:
            page["free_text_regions"].append({
                "id": slug(item.get("id"), f"text_{index + 1}"),
                "label": str(item.get("label") or "free text").strip(),
                "bbox": region,
            })
    return page


def attach_extraction(page: dict[str, Any], raw: dict[str, Any], model: str) -> dict[str, Any]:
    """Attach one model's readings to a normalized page structure."""
    field_map = {field["id"]: field for field in page["metadata_fields"]}
    for reading in raw.get("metadata") or []:
        target = field_map.get(slug(reading.get("field_id"), ""))
        if target is None:
            continue
        target.setdefault("readings", []).append(_reading(reading, model, target["bbox"]))

    text_map = {item["id"]: item for item in page["free_text_regions"]}
    for reading in raw.get("free_text") or []:
        target = text_map.get(slug(reading.get("region_id"), ""))
        if target is None:
            continue
        target.setdefault("readings", []).append(_reading(reading, model, target["bbox"]))

    raw_tables = {slug(table.get("table_id"), ""): table
                  for table in raw.get("tables") or []}
    for table in page["tables"]:
        source = raw_tables.get(table["id"])
        if source is None and len(page["tables"]) == len(raw.get("tables") or []):
            source = (raw.get("tables") or [])[page["tables"].index(table)]
        if not source:
            continue
        ncols = len(table["columns"])
        for row_index, raw_row in enumerate(source.get("rows") or []):
            region = row_bbox(raw_row.get("bbox"))
            if region is None:
                continue
            values = list(raw_row.get("values") or [])[:ncols]
            values += [None] * (ncols - len(values))
            confidences = list(raw_row.get("confidences") or [])[:ncols]
            confidences += [0.0] * (ncols - len(confidences))
            illegible = {int(x) for x in (raw_row.get("illegible_columns") or [])
                         if str(x).lstrip("-").isdigit()}
            row_id = str(raw_row.get("row_id") or f"y{region[1]:.4f}")
            fingerprint = _row_key_fingerprint(table["columns"], values)
            # Match rows by printed identifiers before geometry. Model bbox
            # scales can drift on dense pages even when their literal keys do
            # not; no reference/golden value participates in this join.
            row = _match_row(table["rows"], row_id, region, model, fingerprint)
            if row is None:
                unique_row_id = row_id
                used_row_ids = {existing["id"] for existing in table["rows"]}
                suffix = 2
                while unique_row_id in used_row_ids:
                    unique_row_id = f"{row_id}__{suffix}"
                    suffix += 1
                row = {"id": unique_row_id, "bbox": region,
                       "key_fingerprint": fingerprint, "cells": []}
                for col_index, column in enumerate(table["columns"]):
                    row["cells"].append({
                        "column_id": column["id"],
                        "bbox": [column["x0"], region[1], column["x1"], region[3]],
                        "readings": [],
                    })
                table["rows"].append(row)
            for col_index, value in enumerate(values):
                cell = row["cells"][col_index]
                try:
                    confidence = max(0.0, min(1.0, float(confidences[col_index])))
                except (TypeError, ValueError):
                    confidence = 0.0
                cell["readings"].append({
                    "model": model,
                    "value": None if value is None else str(value).strip(),
                    "confidence": round(confidence, 4),
                    "illegible": col_index in illegible,
                    "bbox": cell["bbox"],
                })
    return page


def _reading(raw, model, default_bbox):
    try:
        confidence = max(0.0, min(1.0, float(raw.get("confidence") or 0)))
    except (TypeError, ValueError):
        confidence = 0.0
    return {
        "model": model,
        "value": None if raw.get("value") is None else str(raw.get("value")).strip(),
        "confidence": round(confidence, 4),
        "illegible": bool(raw.get("illegible")),
        "bbox": bbox(raw.get("bbox")) or default_bbox,
    }


def _row_key_fingerprint(columns, values):
    key_parts = []
    stop_kinds = {"species", "local_name", "free_text", "decimal", "number",
                  "measurement", "temperature", "percent"}
    for column, value in zip(columns, values):
        if str(column.get("value_kind") or "").casefold() in stop_kinds:
            break
        part = re.sub(r"[^0-9a-z]+", "", str(value or "").casefold())
        if part:
            key_parts.append(part)
    fingerprint = "".join(key_parts)
    return fingerprint if len(fingerprint) >= 2 and any(char.isdigit() for char in fingerprint) else None


def _match_row(rows, row_id, region, model=None, fingerprint=None):
    # A row can receive at most one reading from a given model. This prevents a
    # duplicate emission within one response from collapsing onto real data.
    available = [row for row in rows
                 if not model or not any(
                     reading.get("model") == model
                     for cell in row.get("cells") or []
                     for reading in cell.get("readings") or [])]
    keyed = [row for row in available
             if fingerprint and row.get("key_fingerprint") == fingerprint]
    if len(keyed) == 1:
        return keyed[0]
    labelled = [row for row in available if row_id and row["id"] == row_id]
    if len(labelled) == 1:
        return labelled[0]

    cy = (region[1] + region[3]) / 2
    height = max(1e-6, region[3] - region[1])
    strong = []
    for row in available:
        other_height = max(1e-6, row["bbox"][3] - row["bbox"][1])
        ry = (row["bbox"][1] + row["bbox"][3]) / 2
        overlap = max(0.0, min(region[3], row["bbox"][3])
                      - max(region[1], row["bbox"][1]))
        overlap_fraction = overlap / min(height, other_height)
        if (overlap_fraction >= 0.65
                and abs(cy - ry) <= max(0.006, min(height, other_height) * .5)):
            strong.append((abs(cy - ry), row))
    if strong:
        return min(strong, key=lambda pair: pair[0])[1]

    # Outside strong geometric agreement, a literal key remains a hard anchor:
    # non-consecutive IDs must not be merged into a neighbouring row.
    if row_id and not row_id.casefold().startswith("y_"):
        return None
    close = []
    for row in available:
        ry = (row["bbox"][1] + row["bbox"][3]) / 2
        overlap = max(0.0, min(region[3], row["bbox"][3]) - max(region[1], row["bbox"][1]))
        if overlap > 0 or abs(cy - ry) < 0.012:
            close.append((abs(cy - ry), row))
    return min(close, default=(None, None), key=lambda pair: pair[0])[1]


def resolve(document: dict[str, Any]) -> dict[str, Any]:
    """Resolve exact agreement; preserve alternatives on disagreement."""
    models = document.get("models") or []
    for page in document["pages"]:
        for field in page.get("metadata_fields") or []:
            _order_model_readings(field, models)
            _resolve_item(field)
        for item in page.get("free_text_regions") or []:
            _order_model_readings(item, models)
            _resolve_item(item)
        for table in page["tables"]:
            table["rows"].sort(key=lambda row: row["bbox"][1])
            for row in table["rows"]:
                for cell in row["cells"]:
                    _order_model_readings(cell, models)
                    _resolve_item(cell)
            _repair_compound_identifier_columns(table)
            _flag_sparse_structural_rows(table)
    return document


def _repair_compound_identifier_columns(table):
    """Split a reader's consistently merged numeric ID + one-letter code."""
    columns = table.get("columns") or []
    rows = table.get("rows") or []
    for index in range(len(columns) - 1):
        source_column, target_column = columns[index:index + 2]
        source_kind = str(source_column.get("value_kind") or "").casefold()
        target_kind = str(target_column.get("value_kind") or "").casefold()
        if source_kind not in {"identifier", "integer"}:
            continue
        if target_kind not in {"identifier", "categorical_code"}:
            continue
        pairs = [(row["cells"][index], row["cells"][index + 1]) for row in rows
                 if len(row.get("cells") or []) > index + 1]
        models = {reading.get("model") for source, target in pairs
                  for cell in (source, target) for reading in cell.get("readings") or []}
        affected = set()
        for model in models:
            observations = []
            for source, target in pairs:
                source_reading = next((item for item in source.get("readings") or []
                                       if item.get("model") == model), None)
                target_reading = next((item for item in target.get("readings") or []
                                       if item.get("model") == model), None)
                if source_reading and norm_value(source_reading.get("value")):
                    observations.append((source, target, source_reading, target_reading))
            if len(observations) < 8 or any(
                    target_reading and norm_value(target_reading.get("value"))
                    for _source, _target, _source_reading, target_reading in observations):
                continue
            matches = [re.fullmatch(r"(\d+)([A-Za-z])", str(reading.get("value")).strip())
                       for _source, _target, reading, _target_reading in observations]
            if sum(match is not None for match in matches) / len(observations) < .9:
                continue
            for (source, target, source_reading, target_reading), match in zip(observations, matches):
                if match is None or target_reading is None:
                    continue
                source_reading["value"] = match.group(1)
                target_reading.update({"value": match.group(2),
                                       "confidence": source_reading.get("confidence", 0),
                                       "illegible": source_reading.get("illegible", False),
                                       "missing": False, "bbox": target.get("bbox")})
                affected.add((id(source), id(target)))
                provenance = (f"split {model} compound identifier into "
                              f"{source_column.get('label')} and {target_column.get('label')} "
                              "after model-wide empty-column check")
                source["layout_repair"] = provenance
                target["layout_repair"] = provenance
        if affected:
            for source, target in pairs:
                if (id(source), id(target)) in affected:
                    _resolve_item(source)
                    _resolve_item(target)


def _flag_sparse_structural_rows(table):
    """Queue likely row shifts while retaining every literal value."""
    columns = table.get("columns") or []
    descriptor_kinds = {"species", "text", "free_text", "categorical"}
    measurement_kinds = {"decimal", "number", "measurement", "count",
                         "temperature", "percent"}
    descriptors = [index for index, column in enumerate(columns)
                   if str(column.get("value_kind") or "").casefold() in descriptor_kinds]
    measurements = [index for index, column in enumerate(columns)
                    if str(column.get("value_kind") or "").casefold() in measurement_kinds]
    if not descriptors or not measurements:
        return
    for row in table.get("rows") or []:
        cells = row.get("cells") or []
        missing = [index for index in descriptors
                   if index < len(cells) and not norm_value(cells[index].get("value"))]
        populated = [index for index in measurements
                     if index < len(cells) and norm_value(cells[index].get("value"))]
        if not missing or not populated or min(populated) <= min(missing):
            continue
        reason = ("measurement present after an empty descriptor; possible row or column shift; "
                  "literal value retained")
        for index in populated:
            cell = cells[index]
            if cell.get("status") == "agreement":
                cell["status"] = "structural_anomaly"
                cell["structural_reason"] = reason


def _order_model_readings(item, models):
    """Make cross-reader coverage loss an explicit disagreement.

    Rows are discovered independently. If only the peer emits one, its reading
    must never slide into primary position and silently become truth.
    """
    if not models:
        return
    by_model = {reading.get("model"): reading for reading in item.get("readings") or []}
    item["readings"] = [
        by_model.get(model, {"model": model, "value": None, "confidence": 0.0,
                             "illegible": True, "missing": True,
                             "bbox": item.get("bbox")})
        for model in models
    ]


def _resolve_item(item):
    readings = item.get("readings") or []
    values = ["" if r.get("value") is None else str(r.get("value")).strip()
              for r in readings]
    normalized = [norm_value(value) for value in values]
    nonempty = [value for value in values if value]
    primary_value = readings[0].get("value") if readings else None
    primary_normalized = normalized[0] if normalized else ""
    alternatives = sorted({value for value, normalized_value in zip(values, normalized)
                           if value and normalized_value != primary_normalized},
                          key=str.casefold)
    if readings and len(set(normalized)) == 1 and not any(r.get("illegible") for r in readings):
        item.update({"value": readings[0].get("value"), "status": "agreement",
                     "confidence": min(r.get("confidence", 0) for r in readings),
                     "alternatives": []})
    elif not nonempty and readings:
        item.update({"value": None, "status": "blank_or_illegible",
                     "confidence": 0.0, "alternatives": []})
    elif len(readings) >= 3:
        counts = Counter(normalized)
        winner, votes = counts.most_common(1)[0]
        if votes > len(readings) / 2:
            # Correlated vision models are not independent voters. Keep the
            # primary literal even when the other two agree. The majority is a
            # review signal, not permission to replace source transcription.
            item.update({"value": primary_value, "status": "majority_after_reread",
                         "confidence": readings[0].get("confidence", 0),
                         "alternatives": alternatives})
            return
        item.update({"value": primary_value,
                     "status": "unresolved_after_reread", "confidence": 0.0,
                     "alternatives": alternatives})
    else:
        # No majority is silently promoted. Targeted reread or human review must
        # resolve it; the primary value (including blank) stays on display.
        item.update({"value": primary_value, "status": "disagreement",
                     "confidence": 0.0,
                     "alternatives": alternatives})


def norm_value(value) -> str:
    text = "" if value is None else " ".join(str(value).strip().split()).casefold()
    if not text:
        return ""
    try:
        number = float(text)
        return str(int(number)) if math.isfinite(number) and number.is_integer() else str(number)
    except ValueError:
        return text


def new_document(source: str, pages: list[dict[str, Any]], models: list[str]) -> dict[str, Any]:
    return {"version": VERSION, "source": source, "models": models, "pages": pages}


def validate(document: dict[str, Any]) -> list[str]:
    errors = []
    if document.get("version") != VERSION:
        errors.append("unsupported version")
    page_numbers = [page.get("page_number") for page in document.get("pages") or []]
    if len(set(page_numbers)) != len(page_numbers):
        errors.append("duplicate page numbers")
    for page in document.get("pages") or []:
        text_ids = [item["id"] for item in page.get("free_text_regions") or []]
        if len(set(text_ids)) != len(text_ids):
            errors.append(f"page {page['page_number']}: duplicate free-text IDs")
        for table in page.get("tables") or []:
            column_ids = [column["id"] for column in table.get("columns") or []]
            if len(set(column_ids)) != len(column_ids):
                errors.append(f"page {page['page_number']} table {table['id']}: duplicate columns")
            row_ids = [row["id"] for row in table.get("rows") or []]
            if len(set(row_ids)) != len(row_ids):
                errors.append(f"page {page['page_number']} table {table['id']}: duplicate rows")
            for row in table.get("rows") or []:
                cell_ids = [cell["column_id"] for cell in row.get("cells") or []]
                if cell_ids != column_ids:
                    errors.append(f"page {page['page_number']} table {table['id']} row {row['id']}: schema drift")
    return errors


def assign_xlsx_coordinates(document: dict[str, Any]) -> dict[str, Any]:
    """Attach the exact page-local Excel coordinate used by ``write_xlsx``."""
    for page in document["pages"]:
        row_cursor = 1
        for field in page.get("metadata_fields") or []:
            field["xlsx_row"], field["xlsx_column"] = row_cursor, 2
            row_cursor += 1
        for item in page.get("free_text_regions") or []:
            item["xlsx_row"], item["xlsx_column"] = row_cursor, 2
            row_cursor += 1
        if (page.get("metadata_fields") or page.get("free_text_regions")):
            row_cursor += 1
        for table in page.get("tables") or []:
            if table["title"]:
                row_cursor += 1
            if any(column.get("parent") for column in table["columns"]):
                row_cursor += 1
            row_cursor += 1  # leaf labels
            for row in table["rows"]:
                for index, item in enumerate(row["cells"], 1):
                    item["xlsx_row"], item["xlsx_column"] = row_cursor, index
                row_cursor += 1
            row_cursor += 2
    return document


def write_xlsx(document: dict[str, Any], destination: str | Path) -> Path:
    assign_xlsx_coordinates(document)
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for page in document["pages"]:
        ws = workbook.create_sheet(f"page{page['page_number']}")
        row_cursor = 1
        for field in page.get("metadata_fields") or []:
            ws.cell(row_cursor, 1).value = field["label"]
            _write_value(ws.cell(row_cursor, 2), field)
            row_cursor += 1
        for item in page.get("free_text_regions") or []:
            label = "Legend" if "legend" in f"{item['id']} {item['label']}".casefold() else "Note"
            ws.cell(row_cursor, 1).value = label
            _write_value(ws.cell(row_cursor, 2), item)
            row_cursor += 1
        if (page.get("metadata_fields") or page.get("free_text_regions")):
            row_cursor += 1
        for table in page.get("tables") or []:
            if table["title"]:
                ws.cell(row_cursor, 1).value = table["title"]
                ws.cell(row_cursor, 1).font = Font(bold=True)
                if len(table["columns"]) > 1:
                    ws.merge_cells(start_row=row_cursor, start_column=1,
                                   end_row=row_cursor, end_column=len(table["columns"]))
                row_cursor += 1
            parent_row = row_cursor
            label_row = row_cursor + 1
            parents = [column.get("parent") for column in table["columns"]]
            if any(parents):
                start = 0
                while start < len(parents):
                    parent = parents[start]
                    end = start
                    while end + 1 < len(parents) and parents[end + 1] == parent:
                        end += 1
                    if parent:
                        ws.cell(parent_row, start + 1).value = parent
                        if end > start:
                            ws.merge_cells(start_row=parent_row, start_column=start + 1,
                                           end_row=parent_row, end_column=end + 1)
                    start = end + 1
                row_cursor += 1
            for index, column in enumerate(table["columns"], 1):
                cell = ws.cell(row_cursor, index)
                cell.value = column["label"]
                cell.font = Font(bold=True)
                cell.alignment = Alignment(wrap_text=True, horizontal="center")
            row_cursor += 1
            for row in table["rows"]:
                for index, item in enumerate(row["cells"], 1):
                    _write_value(ws.cell(row_cursor, index), item)
                row_cursor += 1
            row_cursor += 2
        ws.freeze_panes = "A2"
        for column_index in range(1, ws.max_column + 1):
            letter = get_column_letter(column_index)
            values = (ws.cell(row, column_index).value for row in range(1, ws.max_row + 1))
            ws.column_dimensions[letter].width = min(
                40, max(10, max((len(str(value)) for value in values if value is not None),
                                default=10) + 2))
    workbook.save(destination)
    return Path(destination)


def _write_value(cell, item):
    cell.value = item.get("value")
    status = item.get("status")
    ecology_flags = item.get("ecology_flags") or []
    has_ecology_review = any(flag.get("severity") in ("high", "medium")
                             for flag in ecology_flags)
    is_disagreement = status in {
        "disagreement", "majority_after_reread", "unresolved_after_reread",
        "structural_anomaly"}
    # Human priority is transcription first: a cell that is both disputed and
    # ecologically unusual stays red. Ecology-only cells are orange.
    cell.fill = (RED if is_disagreement else ORANGE if has_ecology_review
                 else GREEN if status == "agreement" else YELLOW)
    readings = item.get("readings") or []
    details = [f"status: {status}", f"bbox: {item.get('bbox')}"]
    if item.get("structural_reason"):
        details.append(item["structural_reason"])
    if item.get("layout_repair"):
        details.append(item["layout_repair"])
    details.extend(f"{r.get('model')}: {r.get('value')!r} (confidence {r.get('confidence')})"
                   for r in readings)
    if item.get("alternatives"):
        details.append(f"alternatives: {item['alternatives']}")
    for flag in ecology_flags:
        details.append(f"ecology {flag.get('severity')}: {flag.get('code')} — "
                       f"{flag.get('message')}")
        if flag.get("proposed_value") is not None:
            details.append(f"ecology suggestion (not applied): {flag['proposed_value']!r}")
    cell.comment = Comment("\n".join(details), "Formidable")


def dump(document: dict[str, Any], destination: str | Path) -> Path:
    Path(destination).write_text(json.dumps(document, indent=2, ensure_ascii=False))
    return Path(destination)

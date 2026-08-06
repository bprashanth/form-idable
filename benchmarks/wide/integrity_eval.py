#!/usr/bin/env python3
"""Strict workbook integrity metrics and adversarial scorer self-tests.

This complements, rather than replaces, ``wide_diff.py``.  The old scorer asks
whether the right tokens occur anywhere.  This module separately asks whether
they occur on the right page and at the right row/column coordinate.

Strict coordinates are meaningful only when the golden itself is a layout
golden.  Many real ``eval_forms`` goldens are content-consensus workbooks, so
the strict metrics are diagnostics there; exact synthetic/canonical goldens
can use them as headline metrics.  The distinction is printed in every report.
"""
from __future__ import annotations

import argparse
import json
import math
import re
import tempfile
from collections import Counter
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


_PAGE_RE = re.compile(r"(?:page|sheet|p)[ _-]*(\d+)$", re.I)


def norm(value: Any) -> str:
    if value is None:
        return ""
    text = " ".join(str(value).strip().split())
    if not text:
        return ""
    try:
        number = float(text)
        return str(int(number)) if math.isfinite(number) and number.is_integer() else str(number)
    except ValueError:
        return text.casefold()


def sheet_key(name: str, index: int) -> str:
    """Normalise common page-sheet spellings without guessing table identity."""
    match = _PAGE_RE.search(name.strip())
    if match:
        return f"page:{int(match.group(1))}"
    if name.strip().casefold() == "v2":
        return "v2"
    return f"name:{name.strip().casefold()}:{index}"


@dataclass
class Grid:
    cells: dict[tuple[str, int, int], str]
    bounds: dict[str, tuple[int, int]]
    sheet_names: dict[str, str]

    @property
    def pages(self) -> set[str]:
        return set(self.bounds)


def load(path: str | Path, *, exclude_consensus_fix: bool = True) -> Grid:
    workbook = openpyxl.load_workbook(path, data_only=True)
    cells: dict[tuple[str, int, int], str] = {}
    bounds: dict[str, tuple[int, int]] = {}
    names: dict[str, str] = {}
    for index, ws in enumerate(workbook.worksheets):
        if exclude_consensus_fix and ws.title.strip().casefold() == "consensus_fix":
            continue
        key = sheet_key(ws.title, index)
        # Duplicate normalised names are kept separate instead of overwritten.
        if key in bounds:
            key = f"{key}:duplicate:{index}"
        names[key] = ws.title
        max_row = max_col = 0
        for row in ws.iter_rows():
            for cell in row:
                value = norm(cell.value)
                if not value:
                    continue
                cells[(key, cell.row, cell.column)] = value
                max_row = max(max_row, cell.row)
                max_col = max(max_col, cell.column)
        bounds[key] = (max_row, max_col)
    return Grid(cells, bounds, names)


def _f1(precision: float, recall: float) -> float:
    return 2 * precision * recall / (precision + recall) if precision + recall else 0.0


def _prf(hits: int, candidate_n: int, golden_n: int) -> dict[str, float]:
    precision = hits / candidate_n if candidate_n else 0.0
    recall = hits / golden_n if golden_n else 0.0
    return {
        "precision": round(precision, 4),
        "recall": round(recall, 4),
        "f1": round(_f1(precision, recall), 4),
    }


def _duplicate_row_excess(grid: Grid) -> int:
    excess = 0
    for page in grid.pages:
        rows: dict[int, list[tuple[int, str]]] = {}
        for (p, row, col), value in grid.cells.items():
            if p == page:
                rows.setdefault(row, []).append((col, value))
        signatures = Counter(tuple(sorted(values)) for values in rows.values() if values)
        excess += sum(count - 1 for count in signatures.values() if count > 1)
    return excess


def _constancy(grid: Grid) -> dict[tuple[str, int], float]:
    columns: dict[tuple[str, int], list[str]] = {}
    for (page, _row, col), value in grid.cells.items():
        columns.setdefault((page, col), []).append(value)
    out = {}
    for key, values in columns.items():
        if len(values) >= 5:
            out[key] = Counter(values).most_common(1)[0][1] / len(values)
    return out


def _oracle_modal_control(grid: Grid) -> dict[str, Any]:
    """Score an oracle-structure filler that reads no pixels.

    It knows every occupied coordinate and the modal value of each page/column,
    so it is intentionally a strong control.  Beating it is necessary evidence
    of reading, not a proposed extraction strategy.
    """
    by_column: dict[tuple[str, int], list[str]] = {}
    for (page, _row, col), value in grid.cells.items():
        by_column.setdefault((page, col), []).append(value)
    modes = {key: Counter(values).most_common(1)[0][0]
             for key, values in by_column.items()}
    null_cells = {
        pos: modes[(pos[0], pos[2])]
        for pos in grid.cells
    }
    exact_hits = sum(grid.cells[pos] == value for pos, value in null_cells.items())
    content_hits = sum((Counter(grid.cells.values()) & Counter(null_cells.values())).values())
    n = len(grid.cells)
    return {
        "description": "oracle occupied positions plus per-page/column modes; reads no pixels",
        "exact_cell_recall": round(exact_hits / n, 4) if n else 0.0,
        "content_anywhere_f1": round(content_hits / n, 4) if n else 0.0,
        "constant_columns": sum(
            len(values) >= 5 and len(set(values)) > 1 for values in by_column.values()
        ),
    }


def _imbalance_metrics(golden: Grid, candidate: Grid) -> dict[str, Any]:
    """Metrics a modal filler cannot game."""
    groups: dict[tuple[str, int], list[tuple[tuple[str, int, int], str]]] = {}
    for pos, value in golden.cells.items():
        groups.setdefault((pos[0], pos[2]), []).append((pos, value))

    offmode_positions = []
    class_recalls = []
    silent_columns = 0
    for key, items in groups.items():
        values = [value for _, value in items]
        mode = Counter(values).most_common(1)[0][0]
        offmode_positions.extend(pos for pos, value in items if value != mode)
        if len(items) >= 3 and not any((page == key[0] and col == key[1])
                                       for page, _row, col in candidate.cells):
            silent_columns += 1
        by_class: dict[str, list[tuple[str, int, int]]] = {}
        for pos, value in items:
            by_class.setdefault(value, []).append(pos)
        for value, positions in by_class.items():
            class_recalls.append(sum(candidate.cells.get(pos) == value for pos in positions)
                                 / len(positions))

    off_hits = sum(candidate.cells.get(pos) == golden.cells[pos] for pos in offmode_positions)
    return {
        "offmode_accuracy_all": round(off_hits / len(offmode_positions), 4)
        if offmode_positions else None,
        "offmode_n": len(offmode_positions),
        "balanced_class_recall": round(sum(class_recalls) / len(class_recalls), 4)
        if class_recalls else None,
        "silent_columns": silent_columns,
    }
def score(golden: str | Path, candidate: str | Path, *,
          golden_kind: str = "content_consensus") -> dict[str, Any]:
    """Score tokens, occupied coordinates, and exact cells independently."""
    g, c = load(golden), load(candidate)
    g_items, c_items = Counter(g.cells.values()), Counter(c.cells.values())
    content_hits = sum((g_items & c_items).values())

    g_coords, c_coords = set(g.cells), set(c.cells)
    position_hits = len(g_coords & c_coords)
    exact_hits = sum(g.cells[pos] == c.cells[pos] for pos in g_coords & c_coords)

    common = g_coords & c_coords
    wrong_value = sum(g.cells[pos] != c.cells[pos] for pos in common)
    false_fill = len(c_coords - g_coords)
    omitted = len(g_coords - c_coords)

    page_hits = len(g.pages & c.pages)
    page_metrics = _prf(page_hits, len(c.pages), len(g.pages))
    content = _prf(content_hits, len(c.cells), len(g.cells))
    position = _prf(position_hits, len(c.cells), len(g.cells))
    exact = _prf(exact_hits, len(c.cells), len(g.cells))

    g_const, c_const = _constancy(g), _constancy(c)
    shared_const = set(g_const) & set(c_const)
    constancy_excess = max(
        (c_const[key] - g_const[key] for key in shared_const), default=0.0)

    g_rows = sum(bound[0] for bound in g.bounds.values())
    c_rows = sum(bound[0] for bound in c.bounds.values())
    g_cols = sum(bound[1] for bound in g.bounds.values())
    c_cols = sum(bound[1] for bound in c.bounds.values())

    null = _oracle_modal_control(g)
    return {
        "golden_kind": golden_kind,
        "strict_metrics_are_headline": golden_kind == "exact_layout",
        "golden_cells": len(g.cells),
        "candidate_cells": len(c.cells),
        "cell_count_ratio": round(len(c.cells) / len(g.cells), 4) if g.cells else None,
        "pages": {**page_metrics, "golden": len(g.pages), "candidate": len(c.pages)},
        "content_anywhere": content,
        "occupied_position": position,
        "exact_cell": exact,
        "class_imbalance_resistant": _imbalance_metrics(g, c),
        "oracle_modal_control": {
            **null,
            "candidate_content_f1_lift": round(content["f1"] - null["content_anywhere_f1"], 4),
            "candidate_exact_recall_lift": round(exact["recall"] - null["exact_cell_recall"], 4),
        },
        "errors": {
            "omitted": omitted,
            "wrong_value_at_occupied_position": wrong_value,
            "false_fill_or_extra_position": false_fill,
            "duplicate_row_excess": _duplicate_row_excess(c),
        },
        "shape": {
            "row_extent_ratio": round(c_rows / g_rows, 4) if g_rows else None,
            "column_extent_ratio": round(c_cols / g_cols, 4) if g_cols else None,
        },
        "constancy": {
            "max_excess_over_golden": round(max(0.0, constancy_excess), 4),
            "candidate_columns_ge_0_98": sum(v >= 0.98 for v in c_const.values()),
            "golden_columns_ge_0_98": sum(v >= 0.98 for v in g_const.values()),
        },
    }


def modal_null(golden: str | Path, destination: str | Path) -> Path:
    """Oracle-structure, per-page/column modal filler that reads no values."""
    src = openpyxl.load_workbook(golden, data_only=True)
    out = openpyxl.Workbook()
    out.remove(out.active)
    for ws in src.worksheets:
        if ws.title.strip().casefold() == "consensus_fix":
            continue
        dst = out.create_sheet(ws.title)
        modes: dict[int, str] = {}
        for col in range(1, ws.max_column + 1):
            values = [norm(ws.cell(row, col).value) for row in range(1, ws.max_row + 1)]
            values = [value for value in values if value]
            modes[col] = Counter(values).most_common(1)[0][0] if values else ""
        for row in range(1, ws.max_row + 1):
            for col in range(1, ws.max_column + 1):
                if norm(ws.cell(row, col).value):
                    dst.cell(row, col).value = modes[col]
    out.save(destination)
    return Path(destination)


def _mutate(golden: str | Path, destination: str | Path, mutation: str) -> Path:
    """Create scorer-integrity fixtures without modifying the source golden."""
    src = openpyxl.load_workbook(golden, data_only=True)
    out = openpyxl.Workbook()
    out.remove(out.active)
    sheets = [ws for ws in src.worksheets if ws.title.strip().casefold() != "consensus_fix"]

    if mutation == "flatten_pages":
        dst = out.create_sheet("v2")
        target_row = 1
        for ws in sheets:
            for row in ws.iter_rows(values_only=True):
                for col, value in enumerate(row, 1):
                    dst.cell(target_row, col).value = value
                target_row += 1
    else:
        for ws in sheets:
            dst = out.create_sheet(ws.title)
            rows = list(ws.iter_rows(values_only=True))
            if mutation == "shuffle_rows" and len(rows) > 1:
                rows = rows[1::2] + rows[::2]
            for row_index, row in enumerate(rows, 1):
                if mutation == "subset_rows" and row_index % 10:
                    continue
                for col, value in enumerate(row, 1):
                    target_col = col + 1 if mutation == "shift_columns" else col
                    if mutation == "swap_columns" and col in (1, 2):
                        target_col = 3 - col
                    target_row = (row_index + 1
                                  if mutation == "partial_row_slip" and row_index > len(rows) // 2
                                  else row_index)
                    if mutation != "blank":
                        if mutation == "format_noise" and value not in (None, ""):
                            value = (f" {value} " if not isinstance(value, (int, float))
                                     else f"{float(value):.1f}")
                        dst.cell(target_row, target_col).value = value
            if mutation == "duplicate_rows" and rows:
                for offset, row in enumerate(rows[: min(3, len(rows))], dst.max_row + 1):
                    for col, value in enumerate(row, 1):
                        dst.cell(offset, col).value = value
            if mutation == "invent_blanks":
                max_row, max_col = max(1, ws.max_row), max(1, ws.max_column)
                added = 0
                for row in range(1, max_row + 1):
                    for col in range(1, max_col + 1):
                        if ws.cell(row, col).value in (None, ""):
                            dst.cell(row, col).value = "FABRICATED"
                            added += 1
                            if added >= 20:
                                break
                    if added >= 20:
                        break
            if mutation == "phantom_rows" and rows:
                for extra in range(1, 6):
                    target = dst.max_row + 1
                    dst.cell(target, 1).value = f"PHANTOM-{extra}"
                    dst.cell(target, 2).value = "INVENTED"
    out.save(destination)
    return Path(destination)


def integrity_test(golden: str | Path) -> dict[str, Any]:
    """Prove the scorer distinguishes common deceptive failure modes."""
    with tempfile.TemporaryDirectory(prefix="formidable_integrity_") as tmp:
        tmp = Path(tmp)
        identity = score(golden, golden, golden_kind="exact_layout")
        cases: dict[str, dict[str, Any]] = {"identity": identity}
        for mutation in ("blank", "subset_rows", "shift_columns", "swap_columns",
                         "shuffle_rows", "partial_row_slip", "flatten_pages",
                         "duplicate_rows", "phantom_rows", "invent_blanks", "format_noise"):
            path = _mutate(golden, tmp / f"{mutation}.xlsx", mutation)
            cases[mutation] = score(golden, path, golden_kind="exact_layout")
        null_path = modal_null(golden, tmp / "modal_null.xlsx")
        cases["modal_null"] = score(golden, null_path, golden_kind="exact_layout")

        checks = {
            "identity_exact": cases["identity"]["exact_cell"]["f1"] == 1.0,
            "blank_detected": cases["blank"]["errors"]["omitted"] > 0,
            "partial_perfect_output_not_flattered": (
                cases["subset_rows"]["exact_cell"]["recall"] <= 0.2
            ),
            "column_shift_detected_despite_same_content": (
                cases["shift_columns"]["content_anywhere"]["f1"] == 1.0
                and cases["shift_columns"]["exact_cell"]["f1"] < 0.5
            ),
            "column_swap_detected_despite_same_content": (
                cases["swap_columns"]["content_anywhere"]["f1"] == 1.0
                and cases["swap_columns"]["exact_cell"]["f1"] < 1.0
            ),
            "row_shuffle_detected_despite_same_content": (
                cases["shuffle_rows"]["content_anywhere"]["f1"] == 1.0
                and cases["shuffle_rows"]["exact_cell"]["f1"] < 0.95
            ),
            "partial_row_slip_detected": cases["partial_row_slip"]["exact_cell"]["f1"] < 0.9,
            "page_flatten_detected_despite_same_content": (
                cases["flatten_pages"]["content_anywhere"]["f1"] == 1.0
                and cases["flatten_pages"]["pages"]["f1"] < 1.0
            ),
            "duplicates_reduce_precision": (
                cases["duplicate_rows"]["content_anywhere"]["precision"] < 1.0
                and cases["duplicate_rows"]["errors"]["duplicate_row_excess"] > 0
            ),
            "phantom_rows_reduce_precision": (
                cases["phantom_rows"]["content_anywhere"]["precision"] < 1.0
                and cases["phantom_rows"]["errors"]["false_fill_or_extra_position"] > 0
            ),
            "invented_blank_fills_detected": (
                cases["invent_blanks"]["errors"]["false_fill_or_extra_position"] > 0
            ),
            "modal_null_is_not_identity": cases["modal_null"]["exact_cell"]["f1"] < 1.0,
            "format_noise_is_normalised": cases["format_noise"]["exact_cell"]["f1"] == 1.0,
        }
        return {"passed": all(checks.values()), "checks": checks, "cases": cases}


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("golden")
    parser.add_argument("candidate", nargs="?")
    parser.add_argument("--golden-kind", choices=("content_consensus", "exact_layout"),
                        default="content_consensus")
    parser.add_argument("--integrity-test", action="store_true")
    args = parser.parse_args()
    if args.integrity_test:
        result = integrity_test(args.golden)
    elif args.candidate:
        result = score(args.golden, args.candidate, golden_kind=args.golden_kind)
    else:
        parser.error("candidate is required unless --integrity-test is used")
    print(json.dumps(result, indent=2))


if __name__ == "__main__":
    main()

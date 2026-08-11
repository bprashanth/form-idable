from pathlib import Path

import openpyxl

import primary_bridge


def _book(path: Path, rows, title="v2"):
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = title
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_row_alignment_preserves_order_and_skips_unmatched_notes(tmp_path: Path):
    canonical = tmp_path / "canonical.xlsx"
    primary = tmp_path / "primary.xlsx"
    _book(canonical, [
        ["Tree No", "Species", "Score"],
        ["1", "Litsea stocksii", "0"],
        ["2", "Persea macrantha", "4"],
        ["3", "Ficus nervosa", "0"],
    ], "page1")
    _book(primary, [
        ["PDF PAGE 1"],
        ["Tree No", "Species", "Score"],
        ["1", "Litsea stocksii", "0"],
        ["transcription note", "ink at margin"],
        ["2", "Persea macrantha", "H"],
        ["3", "Ficus nervosa", "0"],
    ])

    report = primary_bridge.align_workbooks(canonical, primary)
    pairs = [(item["canonical_row"], item["primary_row"])
             for item in report["bindings"]]

    assert pairs == [(1, 2), (2, 3), (3, 5), (4, 6)]
    assert report["mapped_rows"] == 4
    assert report["canonical_row_coverage"] == 1.0


def test_repetitive_rows_require_leading_anchor():
    left = primary_bridge.WorkbookRow("page1", 1, ("10", "species a", "0", "4", "y"))
    correct = primary_bridge.WorkbookRow("v2", 8, ("10", "species a", "0", "4", "y"))
    neighbour = primary_bridge.WorkbookRow("v2", 9, ("11", "species b", "0", "4", "y"))

    assert primary_bridge.row_similarity(left, correct) > 0.95
    assert primary_bridge.row_similarity(left, neighbour) < 0.55


def test_spacer_columns_and_side_by_side_pages_can_share_rows(tmp_path: Path):
    canonical = tmp_path / "canonical.xlsx"
    workbook = openpyxl.Workbook()
    first = workbook.active
    first.title = "page1"
    first.append(["S No", "Species", "Seedlings", "Saplings"])
    first.append(["1)", "Vatt mal", "1", "4"])
    second = workbook.create_sheet("page2")
    second.append(["No", "Species", "Height"])
    second.append(["1", "Ficus nervosa", "42"])
    workbook.save(canonical)

    primary = tmp_path / "primary.xlsx"
    _book(primary, [
        ["S No", "Species", None, "Seedlings", None, "Saplings",
         None, "No", "Species", "Height"],
        ["1", "Vatt mal", None, "1", None, "4",
         None, "1", "Ficus nervosa", "42"],
    ])

    report = primary_bridge.align_workbooks(canonical, primary)
    pairs = [(item["canonical_sheet"], item["canonical_row"],
              item["primary_row"], item["column_offset"])
             for item in report["bindings"]]
    assert pairs == [("page1", 1, 1, 0), ("page1", 2, 2, 0),
                     ("page2", 1, 1, 7), ("page2", 2, 2, 7)]


def test_bind_primary_uses_printed_header_to_cross_spacer_columns(tmp_path: Path):
    canonical = tmp_path / "canonical.xlsx"
    _book(canonical, [
        ["S No", "Species", "Seedlings", "Saplings", "Remarks"],
        ["1)", "Vast mal", "1", "11/9", None],
    ], "page1")
    primary = tmp_path / "primary.xlsx"
    _book(primary, [
        ["S No", "Species", None, "Seedlings", None, "Saplings", None, "Remarks"],
        [1, "Vatt mal", None, 1, None, 4, None, None],
    ])
    cells = [{
        "column_id": name, "bbox": [0, 0, 1, 1], "value": value,
        "readings": [{"model": "peer-a", "value": value},
                     {"model": "peer-b", "value": value}],
        "xlsx_row": 2, "xlsx_column": index,
    } for index, (name, value) in enumerate([
        ("s_no", "1)"), ("species", "Vast mal"), ("seedlings", "1"),
        ("saplings", "11/9"), ("remarks", None)], 1)]
    document = {"models": ["peer-a", "peer-b"], "pages": [{
        "page_number": 1, "metadata_fields": [], "free_text_regions": [],
        "tables": [{
            "columns": [
                {"label": "S No", "parent": None},
                {"label": "Species", "parent": None},
                {"label": "Seedlings", "parent": None},
                {"label": "Saplings", "parent": None},
                {"label": "Remarks", "parent": None},
            ],
            "rows": [{"cells": cells}],
        }],
    }]}

    report = primary_bridge.bind_primary(document, canonical, primary)

    assert report["mapped_items"] == 5
    assert [cell["xlsx_column"] for cell in cells] == [1, 2, 4, 6, 8]
    assert [cell["value"] for cell in cells] == [1, "Vatt mal", 1, 4, None]
    assert cells[1]["status"] == "peer_consensus_disagreement"


def test_checkmark_and_x_are_semantically_equal():
    assert primary_bridge.semantic_value("X") == primary_bridge.semantic_value("✓")
    assert primary_bridge.semantic_value("X") != primary_bridge.semantic_value("Y")

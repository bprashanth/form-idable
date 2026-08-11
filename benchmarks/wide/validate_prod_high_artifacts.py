#!/usr/bin/env python3
"""Cross-validate completed production High artifacts without model calls."""
from __future__ import annotations

import argparse
import json
from pathlib import Path

import fitz
from openpyxl import load_workbook

import primary_bridge
import review_manifest
import run_high_sweep


def valid_bbox(value) -> bool:
    return (isinstance(value, list) and len(value) == 4
            and all(isinstance(part, (int, float)) and 0 <= part <= 1
                    for part in value)
            and value[0] < value[2] and value[1] < value[3])


def workbook_cells_equal(left, right) -> bool:
    if left.sheetnames != right.sheetnames:
        return False
    for name in left.sheetnames:
        a, b = left[name], right[name]
        if a.max_row != b.max_row or a.max_column != b.max_column:
            return False
        for row in range(1, a.max_row + 1):
            for column in range(1, a.max_column + 1):
                if primary_bridge.semantic_value(a.cell(row, column).value) != \
                        primary_bridge.semantic_value(b.cell(row, column).value):
                    return False
    return True


def validate_fixture(name: str, root: Path) -> tuple[dict, list[str]]:
    errors: list[str] = []
    run_dir = root / name
    required = ["output.xlsx", "content_output.xlsx", "canonical.json",
                "review_manifest.json", "analytics.json", "ecology_review.json",
                "crops_manifest.json", "run.json"]
    missing = [filename for filename in required if not (run_dir / filename).exists()]
    if missing:
        return {}, [f"missing artifacts: {', '.join(missing)}"]

    document = json.loads((run_dir / "canonical.json").read_text())
    review = json.loads((run_dir / "review_manifest.json").read_text())
    analytics = json.loads((run_dir / "analytics.json").read_text())
    ecology = json.loads((run_dir / "ecology_review.json").read_text())
    crops = json.loads((run_dir / "crops_manifest.json").read_text())
    run = json.loads((run_dir / "run.json").read_text())
    pdf_pages = fitz.open(run_high_sweep.FORMS / name / "input.pdf").page_count

    errors.extend(review_manifest.validate(review))
    if review.get("version") != "formidable-review-v1":
        errors.append("unsupported review manifest")
    if analytics.get("version") != "formidable-analytics-v1":
        errors.append("unsupported analytics manifest")
    page_counts = {
        "canonical": len(document.get("pages") or []),
        "crops": len(crops.get("pages") or []),
        "analytics": analytics.get("summary", {}).get("pages"),
    }
    if any(count != pdf_pages for count in page_counts.values()):
        errors.append(f"page count mismatch: PDF={pdf_pages}, artifacts={page_counts}")

    content = load_workbook(run_dir / "content_output.xlsx", data_only=True)
    output = load_workbook(run_dir / "output.xlsx", data_only=True)
    output_content_names = [name for name in output.sheetnames
                            if name.casefold() != "ecology_review"]
    if output_content_names != content.sheetnames:
        errors.append("download workbook content sheet order differs from literal workbook")
    else:
        output_content = load_workbook(run_dir / "output.xlsx", data_only=True)
        if "ecology_review" in output_content.sheetnames:
            del output_content["ecology_review"]
        if not workbook_cells_equal(content, output_content):
            errors.append("download workbook changed literal cell content")
    if not output.sheetnames or output.sheetnames[-1] != "ecology_review":
        errors.append("ecology_review is not the final workbook sheet")

    cells = review.get("cells") or []
    by_id = {cell.get("id"): cell for cell in cells}
    if len(by_id) != len(cells):
        errors.append("review cell IDs are not unique")
    for cell in cells:
        page = cell.get("page")
        if not isinstance(page, int) or not 1 <= page <= pdf_pages:
            errors.append(f"{cell.get('id')}: invalid page {page}")
        if not valid_bbox(cell.get("bbox")):
            errors.append(f"{cell.get('id')}: invalid bbox")
        sheet_name = cell.get("xlsx_sheet")
        if not sheet_name:
            if (run.get("route") != "agentic_primary"
                    or cell.get("status") != "unmapped_primary"
                    or cell.get("presented_value") is not None):
                errors.append(f"{cell.get('id')}: invalid unmapped workbook target")
            continue
        if sheet_name not in content.sheetnames:
            errors.append(f"{cell.get('id')}: missing sheet {sheet_name}")
            continue
        row, column = cell.get("xlsx_row"), cell.get("xlsx_column")
        if not isinstance(row, int) or not isinstance(column, int) or row < 1 or column < 1:
            errors.append(f"{cell.get('id')}: invalid workbook coordinate")
            continue
        actual = content[sheet_name].cell(row, column).value
        if primary_bridge.semantic_value(actual) != primary_bridge.semantic_value(
                cell.get("presented_value")):
            errors.append(f"{cell.get('id')}: workbook/review value mismatch")

    red = review.get("views", {}).get("transcription_attention") or []
    red_ids = [item.get("cell_id") for item in red]
    if len(red_ids) != len(set(red_ids)):
        errors.append("duplicate red attention IDs")
    for item in red:
        source = by_id.get(item.get("cell_id"))
        if not source:
            errors.append(f"red target missing review cell: {item.get('cell_id')}")
        elif item.get("page") != source.get("page") or item.get("bbox") != source.get("bbox"):
            errors.append(f"red target geometry differs: {item.get('cell_id')}")

    orange = review.get("views", {}).get("ecology_anomalies") or []
    for item in orange:
        if not valid_bbox(item.get("bbox")):
            errors.append(f"ecology finding {item.get('finding_id')}: invalid bbox")
        if not isinstance(item.get("page"), int) or not 1 <= item["page"] <= pdf_pages:
            errors.append(f"ecology finding {item.get('finding_id')}: invalid page")
        if item.get("xlsx_row") is not None and item.get("xlsx_column") is not None:
            sheet_name = item.get("xlsx_sheet")
            if not sheet_name and f"page{item.get('page')}" in content.sheetnames:
                sheet_name = f"page{item.get('page')}"
            if not sheet_name and len(content.sheetnames) == 1:
                sheet_name = content.sheetnames[0]
            if sheet_name not in content.sheetnames:
                errors.append(
                    f"ecology finding {item.get('finding_id')}: unresolved workbook sheet")

    summary = review.get("summary") or {}
    if summary.get("target_cells_including_blanks") != len(cells):
        errors.append("review target summary differs from cells")
    if summary.get("transcription_review_cells") != len(red):
        errors.append("review red summary differs from view")
    if summary.get("ecology_findings") != len(orange):
        errors.append("review orange summary differs from view")
    analytics_summary = analytics.get("summary") or {}
    if analytics_summary.get("cells") != len(cells):
        errors.append("analytics cell count differs from review")
    if analytics_summary.get("disagreements") != len(red):
        errors.append("analytics disagreement count differs from review")
    actionable = sum(item.get("severity") in {"medium", "high"}
                     for item in ecology.get("findings") or [])
    if actionable != len(orange):
        errors.append("actionable ecology findings differ from orange view")

    return {
        "pages": pdf_pages, "route": run.get("route"), "cells": len(cells),
        "red": len(red), "orange": len(orange), "sheets": content.sheetnames,
    }, errors


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--state", type=Path, required=True)
    args = parser.parse_args()
    state = json.loads(args.state.read_text())
    reports, all_errors = {}, []
    for name in sorted(state["jobs"]):
        report, errors = validate_fixture(name, args.state.parent)
        reports[name] = {**report, "errors": errors}
        all_errors.extend(f"{name}: {error}" for error in errors)
    result = {
        "forms": len(reports), "pages": sum(item.get("pages", 0)
                                              for item in reports.values()),
        "errors": all_errors, "fixtures": reports,
    }
    print(json.dumps(result, indent=2))
    return 1 if all_errors else 0


if __name__ == "__main__":
    raise SystemExit(main())

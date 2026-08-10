#!/usr/bin/env python3
"""Positional, cell-level accuracy for the phenology-style row-table forms.

Why this exists
---------------
`wide_diff.compare()` scores token MULTISETS. It is blind to position, and
worse, it is blind to *fabrication*: a candidate that ignores the handwriting
entirely and stamps each column with its most common value scores ~0.90,
because the true distribution is mode-dominated. See `null_baseline()`.

This scorer keys every data row by its printed Tree No (globally unique across
all six pages of the phenology form) so it is immune to page flattening, and
compares field by field so column drift shows up instead of cancelling out.

Reported:
  value_acc    over golden cells that HAVE a value  <- the number that matters
  blank_prec   of cells the candidate filled where golden is blank, how many
               it correctly left blank (catches invent-a-value behaviour)
  coverage     fraction of golden rows found at all
  per-field accuracy, and a constancy check per field
"""
import sys
from collections import Counter, defaultdict

import openpyxl

# canonical field order; every variant we have seen keeps this order, with
# extra/invented columns appearing at the END of a row, not in the middle.
FIELDS = ["tree_no", "species", "h_m", "gbh_cm", "multistem",
          "lf_flush", "lf_mature", "lf_fallen",
          "fl_buds", "fl_open", "fl_fallen",
          "fr_unripe", "fr_ripe", "fr_fallen", "notes"]
PHENO = FIELDS[5:14]          # the nine hand-filled single-character columns


def _norm(v) -> str:
    """Normalise a cell for comparison: case, whitespace, 4 vs 4.0."""
    if v is None:
        return ""
    s = str(v).strip()
    if not s:
        return ""
    try:                                    # 4.0 -> 4, keeps 118, 18 as-is
        f = float(s)
        return str(int(f)) if f == int(f) else str(f)
    except ValueError:
        return s.casefold()


def records(path: str) -> dict[int, list[str]]:
    """Every data row in every sheet, keyed by Tree No.

    A data row is one whose first cell is an integer. That is what makes this
    robust to the single-sheet flattening production emits: we never rely on
    sheet boundaries, row offsets, or repeated header rows.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    out: dict[int, list[str]] = {}
    for name in wb.sheetnames:
        for row in wb[name].iter_rows(values_only=True):
            if not row or row[0] is None:
                continue
            head = str(row[0]).strip()
            if not head.isdigit():
                continue
            key = int(head)
            vals = [_norm(c) for c in row[:len(FIELDS)]]
            vals += [""] * (len(FIELDS) - len(vals))
            # first sheet wins: the golden's consensus_fix sheet re-lists a few
            # trees, and prod repeats header blocks, but never a whole row.
            out.setdefault(key, vals)
    return out


def score(golden_path: str, cand_path: str) -> dict:
    g, c = records(golden_path), records(cand_path)
    found = [k for k in g if k in c]

    per_field, hits, tot = {}, 0, 0
    false_fill = correct_blank = 0
    for i, f in enumerate(FIELDS):
        n = ok = 0
        for k in found:
            gv, cv = g[k][i], c[k][i]
            if gv:                       # golden has a value here
                n += 1
                ok += (gv == cv)
            else:                        # golden is blank
                if cv:
                    false_fill += 1
                else:
                    correct_blank += 1
        per_field[f] = (ok / n) if n else None
        hits += ok
        tot += n

    # constancy: is the candidate just stamping one value down a column?
    const = {}
    for i, f in enumerate(FIELDS):
        vals = [c[k][i] for k in found if c[k][i]]
        if vals:
            v, n = Counter(vals).most_common(1)[0]
            const[f] = (v, n / len(vals))

    return {
        "golden_rows": len(g),
        "cand_rows": len(c),
        "coverage": round(len(found) / len(g), 3) if g else 0.0,
        "value_acc": round(hits / tot, 3) if tot else 0.0,
        "value_n": tot,
        "pheno_acc": round(_mean([per_field[f] for f in PHENO]), 3),
        "blank_prec": round(correct_blank / (correct_blank + false_fill), 3)
                      if (correct_blank + false_fill) else None,
        "false_fill": false_fill,
        "per_field": {k: (round(v, 3) if v is not None else None)
                      for k, v in per_field.items()},
        "constancy": const,
    }


def _mean(xs):
    xs = [x for x in xs if x is not None]
    return sum(xs) / len(xs) if xs else 0.0


def null_baseline(golden_path: str) -> dict:
    """The control every candidate must beat.

    Fills each field with that field's most common value in the GOLDEN, i.e.
    reads nothing at all. Whatever this scores is the floor; a model at or
    below it has demonstrated no reading ability on this form.
    """
    g = records(golden_path)
    modes = []
    for i in range(len(FIELDS)):
        vals = [r[i] for r in g.values() if r[i]]
        modes.append(Counter(vals).most_common(1)[0][0] if vals else "")
    rows = {k: list(modes) for k in g}
    for k in rows:                       # keep the printed key honest
        rows[k][0] = g[k][0]
    hits = tot = 0
    per_field = {}
    for i, f in enumerate(FIELDS):
        n = ok = 0
        for k in g:
            if g[k][i]:
                n += 1
                ok += (g[k][i] == rows[k][i])
        per_field[f] = round(ok / n, 3) if n else None
        hits += ok
        tot += n
    return {"value_acc": round(hits / tot, 3), "value_n": tot,
            "pheno_acc": round(_mean([per_field[f] for f in PHENO]), 3),
            "modes": {f: modes[i] for i, f in enumerate(FIELDS)},
            "per_field": per_field}


if __name__ == "__main__":
    if sys.argv[1] == "--null":
        import json
        print(json.dumps(null_baseline(sys.argv[2]), indent=2))
    else:
        import json
        print(json.dumps(score(sys.argv[1], sys.argv[2]), indent=2))

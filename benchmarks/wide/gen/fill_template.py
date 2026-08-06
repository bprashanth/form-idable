#!/usr/bin/env python3
"""Fill a REAL blank field-datasheet template with synthetic handwriting.

Why this exists: our synthetic corpus so far uses layouts *I* invented, so both
the model and the template-matcher risk looking good on structures they have
effectively memorised. Real blank templates downloaded from the web supply
authentic structural diversity that nobody on this project designed.

The trick that makes the goldens exact and avoids a CV-parameter treadmill:
a blank template PDF is a VECTOR document. PyMuPDF gives us

  * the printed label text with exact positions  (page.get_text)
  * the printed rules as exact line segments      (page.get_drawings)

so the grid and the labels are read off the file rather than inferred from
pixels. Nothing to tune. We then fill only the cells that are EMPTY in the
blank — which needs no semantic understanding of the form at all — and the
golden is (printed labels, exactly) + (values we wrote, exactly).

Usage:
  python3 fill_template.py <blank.pdf> <out_dir> [--seed N] [--density 0.6]
                           [--page N] [--hard]
"""
import argparse, json, random, sys
from pathlib import Path

import fitz
import openpyxl
from PIL import Image

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))
import formgen2 as fg                                     # noqa: E402

MAX_DIM = 2000


# ── structure straight out of the PDF (no CV, no thresholds) ──────
def extract_structure(pdf: Path, page_no=0):
    doc = fitz.open(str(pdf))
    page = doc[page_no]
    R = page.rect
    words = [{"text": w[4], "bbox": (w[0], w[1], w[2], w[3])}
             for w in page.get_text("words")]
    # Real templates draw rules as THIN FILLED RECTANGLES, and one visual rule
    # is often several rectangles laid end to end with small gaps. So: reduce
    # each thin rect to a single segment along its long axis, then merge
    # collinear, nearly-touching segments into runs.
    hseg, vseg = [], []

    def add(x0, y0, x1, y1, thick=0.0):
        if abs(y1 - y0) <= max(2.5, thick) and abs(x1 - x0) > 4:
            hseg.append((min(x0, x1), max(x0, x1), (y0 + y1) / 2))
        elif abs(x1 - x0) <= max(2.5, thick) and abs(y1 - y0) > 4:
            vseg.append((min(y0, y1), max(y0, y1), (x0 + x1) / 2))

    for d in page.get_drawings():
        for item in d["items"]:
            if item[0] == "l":
                add(item[1].x, item[1].y, item[2].x, item[2].y)
            elif item[0] == "re":
                r = item[1]
                w, h = r.x1 - r.x0, r.y1 - r.y0
                if min(w, h) <= 2.5:                 # a rule, not a box
                    add(r.x0, r.y0, r.x1, r.y1)
                else:                                # a genuine box: 4 sides
                    hseg += [(r.x0, r.x1, r.y0), (r.x0, r.x1, r.y1)]
                    vseg += [(r.y0, r.y1, r.x0), (r.y0, r.y1, r.x1)]
    doc.close()
    return {"rect": (R.x0, R.y0, R.x1, R.y1), "words": words,
            "h": _merge(hseg), "v": _merge(vseg), "page_no": page_no}


def _merge(segs, coord_tol=1.6, gap_tol=4.0):
    """Join collinear segments that touch or nearly touch."""
    if not segs:
        return []
    segs = sorted(segs, key=lambda s: (round(s[2] / coord_tol), s[0]))
    out = []
    for a0, a1, c in segs:
        if out and abs(out[-1][2] - c) <= coord_tol and a0 <= out[-1][1] + gap_tol:
            out[-1] = (out[-1][0], max(out[-1][1], a1), (out[-1][2] + c) / 2)
        else:
            out.append((a0, a1, c))
    return out


def _cluster(vals, tol=2.0):
    vals = sorted(vals)
    out = []
    for v in vals:
        if not out or v - out[-1][-1] > tol:
            out.append([v])
        else:
            out[-1].append(v)
    return [sum(g) / len(g) for g in out]


def build_cells(st, min_w=14, min_h=8):
    """Lattice cells from the rule segments; keep only cells whose four sides
    are actually drawn, so we do not invent cells in whitespace."""
    ys = _cluster([s[2] for s in st["h"]])
    xs = _cluster([s[2] for s in st["v"]])
    if len(ys) < 2 or len(xs) < 2:
        return []

    def h_covers(y, x0, x1):
        return any(abs(s[2] - y) <= 2.5 and s[0] <= x0 + 3 and s[1] >= x1 - 3
                   for s in st["h"])

    def v_covers(x, y0, y1):
        return any(abs(s[2] - x) <= 2.5 and s[0] <= y0 + 3 and s[1] >= y1 - 3
                   for s in st["v"])

    # Build elementary horizontal intervals, then merge through x coordinates
    # whose vertical segment does not actually cross this row. Without this,
    # an unrelated rule elsewhere on the page splits a wide Comments cell into
    # phantom narrow cells because `xs` is page-global.
    cells = []
    for i in range(len(ys) - 1):
        elementary = []
        for j in range(len(xs) - 1):
            y0, y1, x0, x1 = ys[i], ys[i + 1], xs[j], xs[j + 1]
            if (y1 - y0) < min_h:
                continue
            if h_covers(y0, x0, x1) and h_covers(y1, x0, x1):
                elementary.append([j, x0, x1])
        runs = []
        for j, x0, x1 in elementary:
            if (runs and abs(runs[-1][2] - x0) <= 2.0
                    and not v_covers(x0, ys[i], ys[i + 1])):
                runs[-1][2] = x1
            else:
                runs.append([j, x0, x1])
        for j, x0, x1 in runs:
            if x1 - x0 < min_w:
                continue
            # Require 3 of 4 sides overall: both horizontals are known, and at
            # least one vertical permits genuine unruled outer table edges.
            if v_covers(x0, ys[i], ys[i + 1]) or v_covers(x1, ys[i], ys[i + 1]):
                end_col = min(range(j + 1, len(xs)), key=lambda index: abs(xs[index] - x1))
                cells.append({"row": i, "rowspan": 1, "col": j,
                              "colspan": max(1, end_col - j),
                              "bbox": (x0, ys[i], x1, ys[i + 1])})

    # Recover merged faces that span multiple elementary rows. The horizontal-
    # band pass above cannot see a tall header when the intermediate rule stops
    # exactly at that header's edge. Strict four-sided candidates are added only
    # when they do not overlap an existing face; this preserves genuine
    # three-sided outer cells found by the established detector.
    def overlaps(first, second):
        ax0, ay0, ax1, ay1 = first["bbox"]
        bx0, by0, bx1, by1 = second["bbox"]
        return min(ax1, bx1) - max(ax0, bx0) > 2 and min(ay1, by1) - max(ay0, by0) > 2

    strict = []
    for i, y0 in enumerate(ys[:-1]):
        for end_row in range(i + 1, len(ys)):
            y1 = ys[end_row]
            if y1 - y0 < min_h:
                continue
            boundaries = [(j, x) for j, x in enumerate(xs) if v_covers(x, y0, y1)]
            for (j, x0), (end_col, x1) in zip(boundaries, boundaries[1:]):
                if x1 - x0 < min_w:
                    continue
                if not (h_covers(y0, x0, x1) and h_covers(y1, x0, x1)):
                    continue
                internal_h = any(
                    y0 + 2.5 < segment[2] < y1 - 2.5
                    and min(x1, segment[1]) - max(x0, segment[0]) > 4
                    for segment in st["h"])
                internal_v = any(
                    x0 + 2.5 < segment[2] < x1 - 2.5
                    and min(y1, segment[1]) - max(y0, segment[0]) > 4
                    for segment in st["v"])
                if internal_h or internal_v:
                    continue
                strict.append({"row": i, "rowspan": end_row - i,
                               "col": j, "colspan": end_col - j,
                               "bbox": (x0, y0, x1, y1)})
    for candidate in strict:
        exact = next((cell for cell in cells
                      if all(abs(a - b) <= 2 for a, b in
                             zip(cell["bbox"], candidate["bbox"]))), None)
        if exact:
            exact["rowspan"] = candidate["rowspan"]
            exact["colspan"] = candidate["colspan"]
        elif not any(overlaps(candidate, cell) for cell in cells):
            cells.append(candidate)
    page_x0, page_y0, page_x1, page_y1 = st["rect"]
    return [cell for cell in cells
            if cell["bbox"][0] >= page_x0 - 1 and cell["bbox"][1] >= page_y0 - 1
            and cell["bbox"][2] <= page_x1 + 1 and cell["bbox"][3] <= page_y1 + 1]


def cell_text(st, bbox, pad=1.5):
    x0, y0, x1, y1 = bbox
    got = []
    for w in st["words"]:
        wx0, wy0, wx1, wy1 = w["bbox"]
        cx, cy = (wx0 + wx1) / 2, (wy0 + wy1) / 2
        if x0 - pad <= cx <= x1 + pad and y0 - pad <= cy <= y1 + pad:
            got.append((wy0, wx0, w["text"]))
    return " ".join(t for _, _, t in sorted(got))


# ── value generators, chosen per column ───────────────────────────
def _species(rng):    return rng.choice(fg.SPECIES)
def _vernacular(rng): return rng.choice(VERNACULAR)
def _int(rng):        return str(rng.randint(1, 250))
def _dec(rng):        return f"{rng.uniform(0.2, 180):.1f}"
def _code1(rng):      return rng.choice("TSCFNMXYLDR")
def _alpha_code(rng): return "".join(rng.choice("ABCDEFGHJKLMNPRSTUVWY") for _ in range(4))
def _yn(rng):         return rng.choice(["Y", "N"])
def _date(rng):       return f"{rng.randint(1,28):02d}/{rng.randint(1,12):02d}/{rng.randint(19,25)}"
def _time(rng):       return f"{rng.randint(5,18):02d}:{rng.choice([0,10,15,20,30,40,45,50]):02d}"
def _percent(rng):    return str(rng.randint(0, 100))
def _temperature(rng): return f"{rng.uniform(-3, 42):.1f}"
def _ph(rng):         return f"{rng.uniform(3.2, 9.5):.1f}"
def _coordinate(rng): return f"{rng.uniform(-89, 89):.5f}"
def _short(rng):      return rng.choice(fg.REMARKS)
def _person(rng):     return rng.choice(fg.FIRST)

# vernacular names mined from the partner's own transcriptions plus common
# South-Indian tree names — the class of token the corpus was missing entirely
VERNACULAR = ["Kage", "Boothahami", "Icoruu", "Korum", "Kislor", "Rongon",
              "Ulumai", "Selai", "Seengali", "Pusa", "Puli", "Pala", "Padhuva",
              "Mesdhini", "Maaki", "Kadukai", "Punga", "Athi", "Belai", "Eotti",
              "Nooli", "Sandan", "Karimaram", "Vengai", "Thembavu", "Mine",
              "Neem", "Sambrani", "Kanuvai", "Illupai", "Naval", "Vagai"]

COLTYPES = [("species", _species, 0.10), ("vernacular", _vernacular, 0.16),
            ("int", _int, 0.24), ("dec", _dec, 0.16), ("code1", _code1, 0.14),
            ("yn", _yn, 0.06), ("date", _date, 0.05), ("short", _short, 0.05),
            ("person", _person, 0.04)]


# Value kinds a column can hold, by how wide the column actually is. Real form
# designers size a column to its content, so assigning a long species name to a
# 30pt column produces text that overflows into the neighbouring printed label —
# which looks nothing like a real sheet and corrupts the eval image.
WIDTH_BANDS = [
    (38,  ["code1", "yn"]),
    (70,  ["code1", "yn", "int", "date"]),
    (120, ["int", "dec", "date", "person", "vernacular"]),
    (1e9, ["species", "vernacular", "short", "person", "int", "dec"]),
]


def pick_coltypes(cells, ncols, rng):
    """Choose a value kind per column, constrained by that column's width."""
    widths = {}
    for c in cells:
        w = c["bbox"][2] - c["bbox"][0]
        widths.setdefault(c["col"], []).append(w)
    out = []
    for col in range(ncols):
        ws = sorted(widths.get(col, [60]))
        med = ws[len(ws) // 2]
        allowed = next(kinds for lim, kinds in WIDTH_BANDS if med <= lim)
        if col == 0 and med <= 70 and rng.random() < 0.6:
            out.append("int")                      # serial-number column
        else:
            out.append(rng.choice(allowed))
    return out


GEN = {k: f for k, f, _ in COLTYPES}
GEN.update({"time": _time, "percent": _percent, "temperature": _temperature,
            "ph": _ph, "coordinate": _coordinate, "alpha_code": _alpha_code})


def semantic_kind(label, fallback):
    """Infer a plausible value kind from generic field semantics.

    This is deliberately ontology-level, not template-specific: it knows that
    a date holds a date and a scientific-name field holds a taxon, but never
    learns the values or ranges of a benchmark form.
    """
    text = " ".join(str(label).casefold().replace("_", " ").split())
    if not text:
        return fallback
    # Qualifiers beat broad nouns: "species alpha code" is a short code, not
    # a scientific name, and "observer number" is an identifier, not a name.
    if any(phrase in text for phrase in ("species code", "species alpha code",
                                          "taxon code", "alpha code")):
        return "alpha_code"
    if any(mark in text for mark in ("yes/no", " y/n", "", "☐", "check", "tick")):
        return "yn"
    if "mortality" in text and "disease code" not in text:
        return "percent"
    if re_search(text, r"\bas\s+[a-z](?:\s*,\s*[a-z]){2,}"):
        return "code1"
    if (any(word in text for word in ("count", "number", " no.", " no ", "s.no",
                                      "#", "quantity", "isolates", "clumps", "fragments",
                                      "seedling", "sapling"))):
        return "int"
    if (any(word in text for word in ("status", "condition", "habit", "phase", "code",
                                      "seen", "heard", "within", "between", "category",
                                      "class", "grade", "tier"))
            or re_search(text, r"(?:<|>)\s*\d|\b\d+\s*-\s*\d+\b")):
        return "code1"
    if any(word in text for word in ("scientific name", "species", "taxon", "spp")):
        return "species"
    if any(word in text for word in ("local name", "common name", "vernacular")):
        return "vernacular"
    if "date" in text or re_search(text, r"\bdo[csfgt]\b"):
        return "date"
    if "time" in text:
        return "time"
    if any(word in text for word in ("latitude", "longitude", " lat", " long", "gps")):
        return "coordinate"
    if re_search(text, r"\bph\b"):
        return "ph"
    if any(word in text for word in ("temperature", "temp", "°c", "o c")):
        return "temperature"
    if any(word in text for word in ("percent", "%", "cover", "canopy")):
        return "percent"
    if any(word in text for word in ("observer", "collector", "surveyor", "researcher", "name of")):
        return "person"
    if any(word in text for word in ("remark", "note", "comment", "description", "location", "site")):
        return "short"
    if any(word in text for word in ("present", "absence", "yes/no", " y/n", "alive", "survival")):
        return "yn"
    if any(word in text for word in ("depth", "height", "diameter", "dbh", "gbh", "weight",
                                     "length", "width", "moisture", "rainfall", "mass")):
        return "dec"
    if any(word in text for word in ("count", "number", " no.", " no ", "s.no", "#", "quantity",
                                     "plot", "quadrat", "transect", "seedling", "sapling")):
        return "int"
    return fallback


def re_search(text, pattern):
    # Local import keeps the top of this generator lightweight when imported by
    # training scripts that do not call the template filler.
    import re
    return re.search(pattern, text) is not None


def _overlap(a0, a1, b0, b1):
    return max(0.0, min(a1, b1) - max(a0, b0))


def semantic_context(st, cell):
    """Printed field context above a cell and immediately beside it.

    Blank templates often use merged parent headers or labels outside the
    ruled answer box. Restricting inference to text inside the same lattice
    column therefore produced nonsense despite perfect coordinate truth.
    This uses only the blank form's vector text and geometry; no filled value
    or benchmark answer participates.
    """
    x0, y0, x1, y1 = cell["bbox"]
    width, height = x1 - x0, y1 - y0
    candidates = []
    for word in st["words"]:
        wx0, wy0, wx1, wy1 = word["bbox"]
        wwidth, wheight = max(1.0, wx1 - wx0), max(1.0, wy1 - wy0)
        horizontal = _overlap(x0, x1, wx0, wx1)
        vertical = _overlap(y0, y1, wy0, wy1)
        # Column/parent headers. Include modest horizontal expansion so a
        # centred merged header can describe several narrow child columns.
        above = (wy1 <= y0 + 2 and y0 - wy1 <= max(220, height * 8)
                 and wx1 >= x0 - width * 1.25 and wx0 <= x1 + width * 1.25)
        # Field labels printed in the same row immediately to the left/right.
        beside = (vertical / min(height, wheight) >= .25
                  and min(abs(x0 - wx1), abs(wx0 - x1)) <= max(260, width * 4))
        # A word whose centre lies in the cell is the strongest context (some
        # answer cells contain units or a prompt but still have writing room).
        inside = horizontal > 0 and vertical > 0
        if not (above or beside or inside):
            continue
        if inside:
            rank = 0
        elif beside:
            rank = 1 + min(abs(x0 - wx1), abs(wx0 - x1)) / 1000
        else:
            rank = 2 + (y0 - wy1) / 1000
        candidates.append((rank, wy0, wx0, word["text"]))
    # Keep the closest 24 tokens, then restore reading order. This captures
    # nested headings without letting a page title dominate every cell.
    chosen = sorted(candidates)[:24]
    chosen.sort(key=lambda item: (item[1], item[2]))
    seen, words = set(), []
    for _, _, _, word in chosen:
        key = word.casefold()
        if key not in seen:
            words.append(word)
            seen.add(key)
    return " ".join(words)


def nearest_header(st, cell):
    """Closest vector-text header vertically above and overlapping the cell."""
    x0, y0, x1, _ = cell["bbox"]
    width = x1 - x0
    exact, expanded = [], []
    for word in st["words"]:
        wx0, wy0, wx1, wy1 = word["bbox"]
        overlap = _overlap(x0, x1, wx0, wx1)
        direct = overlap / max(1.0, min(width, wx1 - wx0)) >= .2
        nearby = ((wx0 + wx1) / 2 >= x0 - width * 1.25
                  and (wx0 + wx1) / 2 <= x1 + width * 1.25)
        if wy1 <= y0 + 2 and (direct or nearby):
            (exact if direct else expanded).append((wy1, wy0, wx0, word["text"]))
    words = exact or expanded
    if not words:
        return ""
    nearest_y = max(item[0] for item in words)
    chosen = [item for item in words if nearest_y - item[0] <= 25]
    return " ".join(item[3] for item in sorted(chosen, key=lambda item: (item[1], item[2])))


def cell_kinds(st, cells, printed, coltypes):
    """Pick a type per cell from generic vector-text field context."""
    headers = {}
    for cell, text in printed:
        if text.strip():
            headers.setdefault(cell["col"], []).append((cell["row"], text))
    result, contexts = {}, {}
    for cell in cells:
        candidates = [item for item in headers.get(cell["col"], []) if item[0] <= cell["row"]]
        column_label = max(candidates, default=(None, ""), key=lambda item: item[0])[1]
        direct_header = nearest_header(st, cell)
        context = semantic_context(st, cell)
        # The nearest same-column header is more specific than a merged parent
        # or neighbouring label. Only fall back to broader context when it has
        # no recognized generic semantic signal.
        primary = (semantic_kind(direct_header, None)
                   or semantic_kind(column_label, None))
        inferred = primary or semantic_kind(context, coltypes[cell["col"]])
        label = f"{column_label} | {direct_header} | {context}".strip(" |")
        key = (cell["row"], cell["col"])
        contexts[key] = label
        result[key] = inferred
    return result, contexts


def writable_blanks(cells, printed, blanks):
    """Exclude partially empty merged header rows from grid-form filling."""
    from collections import Counter
    printed_by_row = Counter(cell["row"] for cell, text in printed if text.strip())
    blank_by_row = Counter(cell["row"] for cell, _ in blanks)
    rows = sorted(blank_by_row)
    data_like = set()
    for row in rows:
        p, b = printed_by_row[row], blank_by_row[row]
        if p == 0:
            data_like.add(row)
        elif p == 1 and b >= 3:
            texts = [text.strip() for cell, text in printed
                     if cell["row"] == row and text.strip()]
            if texts and texts[0].replace(".", "", 1).isdigit():
                data_like.add(row)
    # Three data-like rows indicate a row-table. Field/questionnaire forms do
    # not have such a run and still need writable boxes beside printed labels.
    longest = run = 0
    for previous, row in zip([None, *rows], rows):
        run = run + 1 if row in data_like and (previous is None or row == previous + 1) else (1 if row in data_like else 0)
        longest = max(longest, run)
    if longest >= 3:
        return [(cell, text) for cell, text in blanks if cell["row"] in data_like]
    return blanks


def fill(pdf: Path, out_dir: Path, seed=0, density=0.6, page_no=0, hard=False):
    rng = random.Random(seed)
    st = extract_structure(pdf, page_no)
    cells = build_cells(st)
    if len(cells) < 8:
        return None, f"only {len(cells)} cells found — not a gridded form page"

    # Render the blank at high resolution; we draw handwriting onto this.
    doc = fitz.open(str(pdf))
    page = doc[page_no]
    scale = min(MAX_DIM / max(page.rect.width, page.rect.height), 4.0)
    pm = page.get_pixmap(matrix=fitz.Matrix(scale, scale))
    img = Image.frombytes("RGB", (pm.width, pm.height), pm.samples).convert("RGBA")
    doc.close()

    writer = fg.Writer(rng, prose=rng.choice(["hybrid", "hybrid", "glyph"]))
    ncols = max(c["col"] for c in cells) + 1
    coltypes = pick_coltypes(cells, ncols, rng)

    # Cells that already contain printed text are labels/headers -> leave them,
    # and record them in the golden. Empty cells are fillable. This needs no
    # semantic understanding of the form.
    printed, blanks = [], []
    for c in cells:
        t = cell_text(st, c["bbox"])
        (printed if t.strip() else blanks).append((c, t))

    golden_printed = [t for _, t in printed if t.strip()]
    kinds, contexts = cell_kinds(st, cells, printed, coltypes)
    rows_written = {}
    truth = {}
    for c, text in printed:
        truth[(c["row"], c["col"])] = {
            "source": "printed", "value": text.strip(), "mark": "text",
            "value_kind": "printed", "semantic_context": text.strip(),
        }
    for c, _ in blanks:
        truth[(c["row"], c["col"])] = {
            "source": "blank", "value": None, "mark": "blank",
            "value_kind": kinds[(c["row"], c["col"])],
            "semantic_context": contexts[(c["row"], c["col"])],
        }
    fillable = writable_blanks(cells, printed, blanks)
    n_fill = int(len(fillable) * density)
    # fill top-down so the sheet looks partially completed, like a real one
    fillable.sort(key=lambda ct: (ct[0]["row"], ct[0]["col"]))
    for c, _ in fillable[:n_fill]:
        if rng.random() > 0.93:
            continue                                   # occasional skipped cell
        kind = kinds[(c["row"], c["col"])]
        val = GEN[kind](rng)
        x0, y0, x1, y1 = [v * scale for v in c["bbox"]]
        box = (x0, y0, x1, y1)
        size = max(14, min(38, (y1 - y0) * 0.62))
        spec = ("text", val)
        if kind in ("int", "dec") and rng.random() < 0.05:
            spec = ("dot",)                            # dot == 0 convention
        elif rng.random() < 0.04:
            spec = ("strike",)                         # struck == no entry
        fg.draw_cell(type("P", (), {"img": img, "d": None, "ptext": lambda *a, **k: None})(),
                     writer, box, spec, size=size)
        g = fg.golden_of(spec)[0]
        truth[(c["row"], c["col"])] = {
            "source": "written" if g is not None else "blank",
            "value": g,
            "mark": spec[0],
            "value_kind": kind,
            "semantic_context": contexts[(c["row"], c["col"])],
        }
        if g is not None:
            rows_written.setdefault(c["row"], []).append((c["col"], g))

    out_dir.mkdir(parents=True, exist_ok=True)
    final = fg.degrade(img, rng, hard=hard)
    tmp = out_dir / "_t.jpg"; final.save(tmp, "JPEG", quality=90)
    d2 = fitz.open()
    r = fitz.Rect(0, 0, final.width * 72 / 180, final.height * 72 / 180)
    d2.new_page(width=r.width, height=r.height).insert_image(r, filename=str(tmp))
    d2.save(out_dir / "input.pdf"); d2.close(); tmp.unlink(missing_ok=True)

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    ws = wb.create_sheet("printed")
    for t in golden_printed:
        ws.append([t])
    ws2 = wb.create_sheet("written")
    for r_i in sorted(rows_written):
        ws2.append([v for _, v in sorted(rows_written[r_i])])
    wb.save(out_dir / "golden.xlsx")

    # A second golden preserves the ruled lattice. The legacy golden above is
    # retained for historical token scores; this one is the layout truth.
    layout = openpyxl.Workbook()
    lws = layout.active
    lws.title = f"page{page_no + 1}"
    for c in cells:
        item = truth[(c["row"], c["col"])]
        if item["value"] is not None:
            lws.cell(c["row"] + 1, c["col"] + 1).value = item["value"]
        if c.get("rowspan", 1) > 1 or c.get("colspan", 1) > 1:
            lws.merge_cells(start_row=c["row"] + 1, start_column=c["col"] + 1,
                            end_row=c["row"] + c.get("rowspan", 1),
                            end_column=c["col"] + c.get("colspan", 1))
    layout.save(out_dir / "layout_golden.xlsx")

    page_x0, page_y0, page_x1, page_y1 = st["rect"]
    page_w, page_h = page_x1 - page_x0, page_y1 - page_y0
    truth_cells = []
    for c in cells:
        x0, y0, x1, y1 = c["bbox"]
        item = truth[(c["row"], c["col"])]
        truth_cells.append({
            "page": page_no + 1,
            "row": c["row"],
            "col": c["col"],
            "rowspan": c.get("rowspan", 1),
            "colspan": c.get("colspan", 1),
            "bbox_points": [round(v, 3) for v in c["bbox"]],
            "bbox_norm": [round((x0 - page_x0) / page_w, 6),
                          round((y0 - page_y0) / page_h, 6),
                          round((x1 - page_x0) / page_w, 6),
                          round((y1 - page_y0) / page_h, 6)],
            **item,
        })
    ground_truth = {
        "version": "formidable-layout-ground-truth-v1",
        "template": pdf.name,
        "page": page_no + 1,
        "page_size_points": [round(page_w, 3), round(page_h, 3)],
        "legibility": (
            "intended values before degradation; degraded hard variants may contain "
            "human-illegible cells and require a sampled ceiling audit"
        ),
        "cells": truth_cells,
    }
    (out_dir / "ground_truth.json").write_text(json.dumps(ground_truth, indent=2))

    meta = {"template": pdf.name, "page": page_no, "seed": seed,
            "density": density, "hard": hard, "cells": len(cells),
            "printed_cells": len(printed), "filled": sum(len(v) for v in rows_written.values()),
            "fallback_coltypes": coltypes,
            "semantic_kind_counts": dict(__import__("collections").Counter(
                item["value_kind"] for item in truth.values())),
            "writer_cohort": writer.cohort,
            "ground_truth": "ground_truth.json",
            "layout_golden": "layout_golden.xlsx"}
    (out_dir / "provenance.md").write_text(
        f"# {out_dir.name}\n\nREAL blank template `{pdf.name}` (page {page_no+1}) "
        f"filled with synthetic handwriting.\nStructure and printed labels read "
        f"directly from the PDF's vector text/line data — exact, no CV "
        f"thresholds.\nGolden is exact by construction: printed labels from the "
        f"file + the values written here.\n\n```json\n{json.dumps(meta, indent=2)}\n```\n")
    return meta, None


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("pdf"); ap.add_argument("out")
    ap.add_argument("--seed", type=int, default=0)
    ap.add_argument("--density", type=float, default=0.6)
    ap.add_argument("--page", type=int, default=0)
    ap.add_argument("--hard", action="store_true")
    a = ap.parse_args()
    m, err = fill(Path(a.pdf), Path(a.out), a.seed, a.density, a.page, a.hard)
    print(json.dumps(m) if m else f"SKIP: {err}")

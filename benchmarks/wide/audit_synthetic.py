#!/usr/bin/env python3
"""Contract audit for exact-layout synthetic form corpora."""
from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter
from pathlib import Path

import openpyxl


def norm(value):
    if value is None:
        return None
    text = str(value).strip()
    try:
        number = float(text)
        return str(int(number)) if math.isfinite(number) and number.is_integer() else str(number)
    except ValueError:
        return text.casefold()


def valid_value(kind, value):
    text = str(value or "").strip()
    if kind in ("printed", "species", "vernacular", "short", "person", "code1"):
        return bool(text)
    if kind == "alpha_code":
        return re.fullmatch(r"[A-Z]{4}", text) is not None
    if kind == "yn":
        return text in ("Y", "N")
    if kind == "date":
        return re.fullmatch(r"\d{2}/\d{2}/\d{2}", text) is not None
    if kind == "time":
        return re.fullmatch(r"(?:[01]\d|2[0-3]):[0-5]\d", text) is not None
    try:
        number = float(text)
    except ValueError:
        return False
    ranges = {"percent": (0, 100), "temperature": (-10, 60),
              "ph": (0, 14), "coordinate": (-180, 180)}
    return kind not in ranges or ranges[kind][0] <= number <= ranges[kind][1]


def audit_form(form_dir: Path):
    errors = []
    truth = json.loads((form_dir / "ground_truth.json").read_text())
    cells = truth.get("cells") or []
    coords = [(cell["page"], cell["row"], cell["col"]) for cell in cells]
    if len(coords) != len(set(coords)):
        errors.append("duplicate cell coordinates")
    workbook = openpyxl.load_workbook(form_dir / "layout_golden.xlsx", data_only=True)
    expected_merges = set()
    occupied = {}
    written = 0
    kinds = Counter()
    for cell in cells:
        box = cell.get("bbox_norm") or []
        if (len(box) != 4 or not all(0 <= value <= 1 for value in box)
                or box[0] >= box[2] or box[1] >= box[3]):
            errors.append(f"invalid bbox at {cell['row']},{cell['col']}")
        expected = cell.get("value")
        sheet = workbook[f"page{cell['page']}"]
        actual = sheet.cell(
            cell["row"] + 1, cell["col"] + 1).value
        if norm(expected) != norm(actual):
            errors.append(f"layout mismatch at {cell['row']},{cell['col']}")
        if cell.get("source") == "written":
            written += 1
            kind = cell.get("value_kind")
            kinds[kind] += 1
            if expected is None or not valid_value(kind, expected):
                errors.append(f"invalid {kind} value at {cell['row']},{cell['col']}: {expected!r}")
        elif cell.get("source") == "blank" and expected is not None:
            errors.append(f"blank has value at {cell['row']},{cell['col']}")
        rowspan, colspan = cell.get("rowspan", 1), cell.get("colspan", 1)
        if (not isinstance(rowspan, int) or not isinstance(colspan, int)
                or rowspan < 1 or colspan < 1):
            errors.append(f"invalid span at {cell['row']},{cell['col']}")
            continue
        for row in range(cell["row"], cell["row"] + rowspan):
            for column in range(cell["col"], cell["col"] + colspan):
                key = (cell["page"], row, column)
                if key in occupied:
                    errors.append(
                        f"overlapping spans at {cell['row']},{cell['col']} and {occupied[key]}")
                occupied[key] = (cell["row"], cell["col"])
        if rowspan > 1 or colspan > 1:
            expected_merges.add(str(openpyxl.utils.get_column_letter(cell["col"] + 1))
                                + str(cell["row"] + 1) + ":"
                                + str(openpyxl.utils.get_column_letter(
                                    cell["col"] + colspan))
                                + str(cell["row"] + rowspan))
    actual_merges = {str(value) for sheet in workbook.worksheets
                     for value in sheet.merged_cells.ranges}
    if expected_merges != actual_merges:
        missing = sorted(expected_merges - actual_merges)
        extra = sorted(actual_merges - expected_merges)
        errors.append(f"merge contract mismatch missing={missing[:3]} extra={extra[:3]}")
    return {"form": form_dir.name, "cells": len(cells), "written": written,
            "kinds": dict(kinds), "errors": errors}


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("corpus")
    parser.add_argument("--output")
    args = parser.parse_args()
    root = Path(args.corpus)
    forms = [audit_form(path) for path in sorted(root.iterdir())
             if path.is_dir() and (path / "ground_truth.json").exists()]
    result = {
        "corpus": str(root), "forms": len(forms),
        "forms_with_writing": sum(form["written"] > 0 for form in forms),
        "cells": sum(form["cells"] for form in forms),
        "written": sum(form["written"] for form in forms),
        "kind_counts": dict(sum((Counter(form["kinds"]) for form in forms), Counter())),
        "errors": sum(len(form["errors"]) for form in forms),
        "empty_forms": [form["form"] for form in forms if not form["written"]],
        "failed_forms": [form for form in forms if form["errors"]],
    }
    text = json.dumps(result, indent=2)
    if args.output:
        Path(args.output).write_text(text + "\n")
    print(text)
    raise SystemExit(1 if result["errors"] else 0)


if __name__ == "__main__":
    main()

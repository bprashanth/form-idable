#!/usr/bin/env python3
"""Known-template oracle: exact vector lattice + page or row-band VLM reading."""
from __future__ import annotations

import argparse
import base64
import json
import os
import sys
import time
import urllib.error
from pathlib import Path

import openpyxl
from openpyxl.comments import Comment
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from PIL import Image, ImageDraw, ImageFont

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))
sys.path.insert(0, str(HERE / "gen"))
import integrity_eval  # noqa: E402
import structured_extract as legacy  # noqa: E402
import structured_pipeline as api  # noqa: E402
import template_value_eval  # noqa: E402
import wide_bench  # noqa: E402
from fill_template import (build_cells, cell_text, extract_structure,  # noqa: E402
                           nearest_header, semantic_context, writable_blanks)


READ_SCHEMA = {
    "type": "object",
    "properties": {"cells": {"type": "array", "items": {
        "type": "object",
        "properties": {
            "cell_id": {"type": "string"},
            "value": {"type": ["string", "null"]},
            "confidence": {"type": "number"},
            "evidence": {"type": "string"},
        },
        "required": ["cell_id", "value", "confidence", "evidence"],
        "additionalProperties": False,
    }}},
    "required": ["cells"], "additionalProperties": False,
}

PROMPT = """Read handwriting from a paper form using the supplied blank-template
cell map. Return exactly one result for every target cell_id and no others.
The context comes only from printed text in the blank form; it is a locator and
semantic hint, never a value. Use null when the target cell has no handwritten
entry or is struck through. A lone dot used as a value is 0; tally marks are
counted; a tick or X is X. Preserve literal multi-value notations. Never fill a
blank from a sequence, distribution, domain expectation, or neighbouring row.
If pixels are ambiguous, return null with evidence `unreadable`.

Targets use zero-based image_index:
{targets}
"""


def _gemini_with_retry(model, prompt, images, schema, attempts=4):
    for attempt in range(attempts):
        try:
            return api.gemini_json(model, prompt, images, schema)
        except urllib.error.HTTPError as error:
            if error.code != 429 and error.code < 500:
                raise
            try:
                detail = error.read().decode("utf-8", "replace")[:500]
            except Exception:
                detail = ""
            if attempt + 1 == attempts:
                print(f"provider HTTP {error.code}; retries exhausted; {detail}", flush=True)
                raise
            delay = min(30, 10 * (attempt + 1))
            print(f"provider HTTP {error.code}; retrying in {delay}s; {detail}", flush=True)
            time.sleep(delay)


def _claude_json(model, prompt, images, schema):
    key = os.environ.get("ANTHROPIC_API_KEY") or os.environ.get("ANTHROPIC_API_KEY_NEW")
    if not key:
        raise RuntimeError("ANTHROPIC_API_KEY is not configured")
    content = []
    for path in images:
        content.append({"type": "image", "source": {
            "type": "base64", "media_type": "image/png",
            "data": base64.b64encode(Path(path).read_bytes()).decode(),
        }})
    content.append({"type": "text", "text": prompt})
    payload = {
        "model": model, "max_tokens": 64000,
        "thinking": {"type": "disabled"},
        "messages": [{"role": "user", "content": content}],
        "output_config": {"format": {"type": "json_schema", "schema": schema}},
    }
    started = time.time()
    response = wide_bench._post(
        "https://api.anthropic.com/v1/messages", payload,
        {"x-api-key": key, "anthropic-version": "2023-06-01"}, timeout=900)
    text = "".join(block.get("text", "") for block in response.get("content", [])
                   if block.get("type") == "text")
    usage = response.get("usage") or {}
    input_tokens, output_tokens = usage.get("input_tokens", 0), usage.get("output_tokens", 0)
    # Sonnet 5 introductory pricing through 2026-08-31: $2/M input, $10/M output.
    cost = input_tokens * 2 / 1_000_000 + output_tokens * 10 / 1_000_000
    return json.loads(text), {
        "model": model, "in_tok": input_tokens, "out_tok": output_tokens,
        "cost_usd": round(cost, 5), "latency_s": round(time.time() - started, 1),
    }


def _model_json(provider, model, prompt, images, schema):
    if provider == "claude":
        return _claude_json(model, prompt, images, schema)
    if provider == "external":
        raise RuntimeError("external provider requires a pre-existing batch JSON and metadata")
    return _gemini_with_retry(model, prompt, images, schema)


def _cell_id(cell):
    return f"r{cell['row']}_c{cell['col']}"


def _targets(st, cells, page_rect):
    width, height = page_rect[2] - page_rect[0], page_rect[3] - page_rect[1]
    out = []
    for cell in cells:
        x0, y0, x1, y1 = cell["bbox"]
        direct = nearest_header(st, cell)
        context = direct or semantic_context(st, cell)
        out.append({
            "cell_id": _cell_id(cell), "row": cell["row"], "col": cell["col"],
            "bbox_1000": [round(1000 * (x0 - page_rect[0]) / width),
                          round(1000 * (y0 - page_rect[1]) / height),
                          round(1000 * (x1 - page_rect[0]) / width),
                          round(1000 * (y1 - page_rect[1]) / height)],
            "context": " ".join(context.split())[:180] or "unlabelled",
        })
    return out


def _row_image(image, scale, row_cells):
    x0 = min(cell["bbox"][0] for cell in row_cells) * scale
    x1 = max(cell["bbox"][2] for cell in row_cells) * scale
    y0 = min(cell["bbox"][1] for cell in row_cells) * scale
    y1 = max(cell["bbox"][3] for cell in row_cells) * scale
    pad_x, pad_y = 8, max(5, (y1 - y0) * .35)
    crop = image.crop((max(0, int(x0 - pad_x)), max(0, int(y0 - pad_y)),
                       min(image.width, int(x1 + pad_x)), min(image.height, int(y1 + pad_y))))
    max_width = 2200
    if crop.width > max_width:
        ratio = max_width / crop.width
        crop = crop.resize((max_width, max(1, int(crop.height * ratio))), Image.LANCZOS)
    label = " | ".join(_cell_id(cell) for cell in row_cells)
    output = Image.new("RGB", (crop.width, crop.height + 28), "white")
    output.paste(crop, (0, 28))
    ImageDraw.Draw(output).text((5, 6), label, fill="black", font=ImageFont.load_default())
    return output


def _write_candidate(st, cells, decisions, destination, sheet_name="page1"):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.sheet_view.showGridLines = False
    thin = Side(style="thin", color="777777")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    widths, heights = {}, {}
    for cell in cells:
        printed = cell_text(st, cell["bbox"]).strip()
        decision = decisions.get(_cell_id(cell)) or {}
        value = printed if printed else decision.get("value")
        target = ws.cell(cell["row"] + 1, cell["col"] + 1)
        if value is not None:
            target.value = value
        target.border = border
        target.alignment = Alignment(vertical="center", wrap_text=True)
        x0, y0, x1, y1 = cell["bbox"]
        colspan, rowspan = cell.get("colspan", 1), cell.get("rowspan", 1)
        per_column = min(34, max(3, (x1 - x0) / 6 / colspan))
        for column in range(cell["col"] + 1, cell["col"] + colspan + 1):
            widths[column] = max(widths.get(column, 0), per_column)
        per_row = min(80, max(12, (y1 - y0) / rowspan))
        for row in range(cell["row"] + 1, cell["row"] + rowspan + 1):
            heights[row] = max(heights.get(row, 0), per_row)
        if printed:
            target.font = Font(bold=True, size=9, color="222222")
            target.fill = PatternFill("solid", fgColor="E7E6E6")
        elif value is not None:
            target.font = Font(size=10, color="1F4E78")
            target.comment = Comment(
                f"model confidence: {decision.get('confidence')}\n"
                f"evidence: {decision.get('evidence')}\n"
                f"source cell: {_cell_id(cell)} bbox={cell['bbox']}", "Formidable")
        if rowspan > 1 or colspan > 1:
            ws.merge_cells(start_row=cell["row"] + 1, start_column=cell["col"] + 1,
                           end_row=cell["row"] + rowspan,
                           end_column=cell["col"] + colspan)
    for column in range(1, ws.max_column + 1):
        letter = openpyxl.utils.get_column_letter(column)
        ws.column_dimensions[letter].width = widths.get(column, 2.2)
    for row, height in heights.items():
        ws.row_dimensions[row].height = height
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    wb.save(destination)


def run(form_dir: Path, template: Path, page_no: int, model: str, mode: str,
        tag: str, rows_per_call=8, targets_per_call=250, provider="gemini",
        input_page=0, output_dir: Path | None = None, prepare_only=False):
    output = output_dir or form_dir / "template_outputs" / tag
    output.mkdir(parents=True, exist_ok=True)
    st = extract_structure(template, page_no)
    cells = build_cells(st)
    printed = [(cell, cell_text(st, cell["bbox"])) for cell in cells]
    blanks = [(cell, text) for cell, text in printed if not text.strip()]
    filled_labels = [(cell, text) for cell, text in printed if text.strip()]
    writable = writable_blanks(cells, filled_labels, blanks)
    writable_cells = [cell for cell, _ in writable]

    blank, _ = legacy._render(template, page_no)
    filled, _ = legacy._render(form_dir / "input.pdf", input_page)
    aligned, did_align = legacy.align(filled.resize(blank.size), blank)
    scale = blank.width / (st["rect"][2] - st["rect"][0])
    targets = _targets(st, writable_cells, st["rect"])
    target_by_id = {target["cell_id"]: target for target in targets}

    batches = []
    if mode == "page":
        for start in range(0, len(targets), targets_per_call):
            batches.append(([aligned], targets[start:start + targets_per_call]))
    else:
        by_row = {}
        for cell in writable_cells:
            by_row.setdefault(cell["row"], []).append(cell)
        rows = sorted(by_row.items())
        for start in range(0, len(rows), rows_per_call):
            chunk = rows[start:start + rows_per_call]
            images = [_row_image(aligned, scale, row_cells) for _, row_cells in chunk]
            items = []
            for image_index, (_, row_cells) in enumerate(chunk):
                for cell in row_cells:
                    item = dict(target_by_id[_cell_id(cell)])
                    item["image_index"] = image_index
                    item.pop("bbox_1000", None)
                    items.append(item)
            batches.append((images, items))

    decisions, calls, duplicates = {}, [], []
    started = time.time()
    if prepare_only:
        for batch_index, (images, items) in enumerate(batches):
            if mode == "page":
                for item in items:
                    item["image_index"] = 0
            (output / f"batch_{batch_index:03d}.targets.json").write_text(
                json.dumps(items, indent=2, ensure_ascii=False))
            for image_index, image in enumerate(images):
                image.save(output / f"batch_{batch_index:03d}__image_{image_index:02d}.png")
        return {"form": form_dir.name, "template": template.name,
                "input_page": input_page + 1, "template_page": page_no + 1,
                "mode": mode, "tag": tag, "batches": len(batches),
                "targets": len(targets), "prepared_only": True}
    for batch_index, (images, items) in enumerate(batches):
        if mode == "page":
            for item in items:
                item["image_index"] = 0
        raw_path = output / f"batch_{batch_index:03d}.json"
        meta_path = output / f"batch_{batch_index:03d}.meta.json"
        targets_path = output / f"batch_{batch_index:03d}.targets.json"
        targets_path.write_text(json.dumps(items, indent=2, ensure_ascii=False))
        image_paths = []
        for image_index, image in enumerate(images):
            image_path = output / f"batch_{batch_index:03d}__image_{image_index:02d}.png"
            if not image_path.exists():
                image.save(image_path)
            image_paths.append(image_path)
        if raw_path.exists() and meta_path.exists():
            raw = json.loads(raw_path.read_text())
            meta = json.loads(meta_path.read_text())
        else:
            raw, meta = _model_json(
                provider, model, PROMPT.format(targets=json.dumps(items, ensure_ascii=False)),
                image_paths, READ_SCHEMA)
            raw_path.write_text(json.dumps(raw, indent=2))
            meta_path.write_text(json.dumps(meta, indent=2))
        calls.append({"batch": batch_index, **meta})
        for item in raw.get("cells") or []:
            cell_id = item.get("cell_id")
            if cell_id in decisions:
                duplicates.append(cell_id)
            decisions[cell_id] = item
        print(f"batch {batch_index + 1}/{len(batches)}: targets={len(items)} "
              f"returned={len(raw.get('cells') or [])}", flush=True)

    expected_ids = {_cell_id(cell) for cell in writable_cells}
    missing = sorted(expected_ids - decisions.keys())
    extra = sorted(decisions.keys() - expected_ids)
    candidate = output / "output.xlsx"
    golden = form_dir / "layout_golden.xlsx"
    # Synthetic form directories contain one extracted page but their exact
    # golden deliberately retains the source template's page number. Real
    # documents must instead use the filled input's page number.
    sheet_number = page_no + 1 if golden.exists() else input_page + 1
    _write_candidate(st, cells, decisions, candidate, f"page{sheet_number}")
    score = (integrity_eval.score(golden, candidate, golden_kind="exact_layout")
             if golden.exists() and input_page == 0 else None)
    literal = (template_value_eval.score(form_dir / "ground_truth.json", decisions)
               if (form_dir / "ground_truth.json").exists() and input_page == 0 else None)
    report = {
        "form": form_dir.name, "template": template.name,
        "page": page_no + 1,  # backwards-compatible alias for template_page
        "input_page": input_page + 1, "template_page": page_no + 1,
        "provider": provider, "model": model, "mode": mode, "tag": tag,
        "aligned": did_align,
        "cells": len(cells), "targets": len(expected_ids), "decisions": len(decisions),
        "missing": missing, "extra": extra, "duplicates": duplicates,
        "calls": calls,
        "cost_usd": (round(sum(call.get("cost_usd") or 0 for call in calls), 5)
                     if all(call.get("cost_usd") is not None for call in calls) else None),
        "cost_complete": all(call.get("cost_usd") is not None for call in calls),
        "latency_s": (round(sum(call.get("latency_s") or 0 for call in calls), 1)
                      if all(call.get("latency_s") is not None for call in calls) else None),
        "latency_complete": all(call.get("latency_s") is not None for call in calls),
        "wall_s": round(time.time() - started, 1),
        "literal": literal, "integrity": score,
    }
    (output / "run.json").write_text(json.dumps(report, indent=2))
    return report


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("form")
    parser.add_argument("--template", required=True)
    parser.add_argument("--page", type=int, default=0)
    parser.add_argument("--input-page", type=int, default=0,
                        help="zero-based page in the filled input PDF")
    parser.add_argument("--model", default="gemini-3.6-flash")
    parser.add_argument("--provider", choices=("gemini", "claude", "external"),
                        default="gemini")
    parser.add_argument("--mode", choices=("page", "bands"), default="bands")
    parser.add_argument("--tag")
    parser.add_argument("--rows-per-call", type=int, default=8)
    parser.add_argument("--targets-per-call", type=int, default=250)
    parser.add_argument("--prepare-only", action="store_true",
                        help="save target maps and images without a provider call")
    args = parser.parse_args()
    tag = args.tag or f"{args.model}__{args.mode}"
    report = run(Path(args.form), Path(args.template), args.page, args.model,
                 args.mode, tag, args.rows_per_call, args.targets_per_call, args.provider,
                 args.input_page, prepare_only=args.prepare_only)
    if args.prepare_only:
        print(json.dumps(report, indent=2))
    else:
        print(json.dumps({key: report[key] for key in
                           ("form", "model", "mode", "aligned", "targets", "decisions",
                           "missing", "extra", "cost_usd", "latency_s", "literal",
                           "integrity")}, indent=2))


if __name__ == "__main__":
    main()

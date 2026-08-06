#!/usr/bin/env python3
"""Cross-converter consensus for golden building.

Several independent models transcribe the same form. The scorer compares token
MULTISETS, so consensus is computed the same way: for every token, take the
count each converter produced and keep the median. That simultaneously fixes
one model hallucinating extra copies and another dropping some.

Two outputs:
  report  — what the base transcription disagrees with the majority about,
            so a human/strong model only has to adjudicate those cells
  build   — write golden.xlsx = base structure + consensus corrections

Usage:
  python3 consensus.py report <form_dir> [--base <tag>]
  python3 consensus.py build  <form_dir> [--base <tag>] [--apply fixes.json]
"""
import json, statistics, sys
from collections import Counter
from pathlib import Path

HERE = Path(__file__).parent
sys.path.insert(0, str(HERE))
GSHEP = Path.home() / "src/github.com/bprashanth/good-shepherd/agents/formidable"
sys.path.insert(0, str(GSHEP))
import xlsx_diff                                    # noqa: E402
import openpyxl                                     # noqa: E402
import wide_bench                                   # noqa: E402

# Converters used to build goldens. Anything listed here is CONTAMINATED as an
# eval model and must be excluded from (or flagged in) the leaderboard.
CONVERTER_TAGS = [
    "gemini__gemini-3.6-flash__pertile",
    "openrouter__qwen_qwen3-vl-32b-instruct__pertile",
    "openrouter__qwen_qwen3-vl-235b-a22b-instruct__pertile",
    "openrouter__z-ai_glm-4.6v__pertile",
    "codex__cli__agentic",
]
DEFAULT_BASE = "gemini__gemini-3.6-flash__pertile"


def _tokens_from_text(text):
    cells = []
    for line in text.splitlines():
        if line.strip().lower().startswith("### page"):
            continue
        cells.append(line)
    return xlsx_diff._atoms(cells)


def _tokens_from_xlsx(path):
    return xlsx_diff._atoms(xlsx_diff._cells(str(path)))


def load_converters(form_dir: Path):
    """-> {tag: (nums, words)} for every converter that produced output."""
    out = {}
    od = form_dir / "outputs"
    for tag in CONVERTER_TAGS:
        txt, xls = od / f"{tag}.txt", od / f"{tag}.xlsx"
        if txt.exists():
            out[tag] = _tokens_from_text(txt.read_text())
        elif xls.exists():
            out[tag] = _tokens_from_xlsx(xls)
    return out


def consensus_counts(per_conv):
    """Median count per token across converters (nums and words separately)."""
    res = {}
    for idx, kind in enumerate(("nums", "words")):
        counters = [Counter(v[idx]) for v in per_conv.values()]
        keys = set().union(*[set(c) for c in counters]) if counters else set()
        merged = Counter()
        for k in keys:
            counts = sorted(c[k] for c in counters)
            med = int(statistics.median(counts))
            if med:
                merged[k] = med
        res[kind] = merged
    return res


def agreement(per_conv, cons):
    """F1 of each converter against the consensus multiset.

    Must be F1, not recall: a converter that duplicates or hallucinates tokens
    still reproduces every consensus token, so recall alone would crown the
    most over-producing model as the base (this happened on eval_01, where a
    base emitting 589 words against a 253-word consensus was selected)."""
    out = {}
    for tag, (n, w) in per_conv.items():
        cn, cw = Counter(n), Counter(w)
        matched = (sum((cons["nums"] & cn).values())
                   + sum((cons["words"] & cw).values()))
        c_tot = sum(cn.values()) + sum(cw.values())
        g_tot = sum(cons["nums"].values()) + sum(cons["words"].values())
        rec = matched / g_tot if g_tot else 0.0
        pre = matched / c_tot if c_tot else 0.0
        out[tag] = round(2 * rec * pre / (rec + pre), 4) if (rec + pre) else 0.0
    return out


def best_base(per_conv, cons, prefer_text=True):
    """Best-F1 converter — the most faithful structural skeleton."""
    ag = agreement(per_conv, cons)
    return max(ag, key=lambda t: ag[t])


def report(form_dir: Path, base_tag=DEFAULT_BASE):
    per_conv = load_converters(form_dir)
    if len(per_conv) < 3:
        print(f"{form_dir.name}: only {len(per_conv)} converters — need >=3")
        return
    cons = consensus_counts(per_conv)
    if base_tag not in per_conv:
        base_tag = next(iter(per_conv))
    base = {"nums": Counter(per_conv[base_tag][0]),
            "words": Counter(per_conv[base_tag][1])}
    print(f"\n=== {form_dir.name}  ({len(per_conv)} converters, base={base_tag})")
    for kind in ("nums", "words"):
        missing = cons[kind] - base[kind]      # majority has, base lacks
        extra = base[kind] - cons[kind]        # base has, majority lacks
        tot = sum(cons[kind].values())
        print(f"  {kind}: consensus {tot} tokens | base missing "
              f"{sum(missing.values())} | base extra {sum(extra.values())}")
        if missing:
            print(f"    MISSING: {', '.join(str(k) for k in sorted(missing)[:40])}"
                  + (" …" if len(missing) > 40 else ""))
        if extra:
            print(f"    EXTRA:   {', '.join(str(k) for k in sorted(extra)[:40])}"
                  + (" …" if len(extra) > 40 else ""))
    agree = {}
    for tag, (n, w) in per_conv.items():
        cn, cw = Counter(n), Counter(w)
        m = sum((cons["nums"] & cn).values()) + sum((cons["words"] & cw).values())
        t = sum(cons["nums"].values()) + sum(cons["words"].values())
        agree[tag] = round(m / t, 3) if t else 0
    print("  per-converter agreement with consensus:")
    for tag, v in sorted(agree.items(), key=lambda kv: -kv[1]):
        print(f"    {v:.3f}  {tag}")


def build(form_dir: Path, base_tag=DEFAULT_BASE, apply=None):
    """golden.xlsx = base transcription structure, plus a `consensus_fix` sheet
    holding majority tokens the base dropped (so the golden's token multiset
    matches the consensus), minus tokens only the base invented."""
    # Never clobber a hand-adjudicated golden: GOLDEN_NOTES.md marks one.
    if (form_dir / "GOLDEN_NOTES.md").exists() and not (apply or "").endswith("force"):
        print(f"{form_dir.name}: hand-adjudicated golden present — skipping "
              f"(delete GOLDEN_NOTES.md to rebuild)")
        return
    per_conv = load_converters(form_dir)
    if len(per_conv) < 3:
        raise SystemExit(f"{form_dir.name}: need >=3 converters")
    cons = consensus_counts(per_conv)
    od = form_dir / "outputs"
    if base_tag == "auto":
        ag = agreement(per_conv, cons)
        # rank by agreement, but only among converters with a .txt skeleton
        cands = [t for t in per_conv if (od / f"{t}.txt").exists()] or list(per_conv)
        base_tag = max(cands, key=lambda t: ag[t])
    base_txt = od / f"{base_tag}.txt"
    if not base_txt.exists():
        # codex only emits xlsx — flatten it into page-marked CSV lines
        wbb = openpyxl.load_workbook(od / f"{base_tag}.xlsx", data_only=True)
        ls = []
        for i, ws in enumerate(wbb.worksheets, 1):
            ls.append(f"### PAGE {i}")
            for row in ws.iter_rows(values_only=True):
                cells = ["" if c is None else str(c) for c in row]
                while cells and cells[-1] == "":
                    cells.pop()
                if cells:
                    ls.append(",".join(f'"{c}"' if ("," in c or '"' in c) else c
                                       for c in (x.replace('"', "'") for x in cells)))
        text = "\n".join(ls)
    else:
        text = base_txt.read_text()

    manual = json.loads(Path(apply).read_text()) if apply else {}
    drop = set(str(x) for x in manual.get("drop", []))
    add = [str(x) for x in manual.get("add", [])]

    lines = [l for l in text.splitlines()
             if not any(d and d in l for d in drop)]

    # Walk the base cells and keep a cell only while the consensus still has
    # budget for its tokens. This makes the golden's token multiset EQUAL the
    # consensus: base-only over-production is dropped, and whatever consensus
    # tokens the base never produced are appended in `consensus_fix`.
    budget = {"nums": Counter(cons["nums"]), "words": Counter(cons["words"])}

    def keep_cell(val):
        if val is None or not str(val).strip():
            return True
        n, w = xlsx_diff._atoms([str(val)])
        if not n and not w:
            return True
        for t in n:
            if budget["nums"][t] <= 0:
                return False
        for t in w:
            if budget["words"][t] <= 0:
                return False
        for t in n:
            budget["nums"][t] -= 1
        for t in w:
            budget["words"][t] -= 1
        return True

    wb = openpyxl.Workbook(); wb.remove(wb.active)
    cur, rows = "page1", []
    import csv as _csv
    dropped = 0

    def flush():
        if rows:
            ws = wb.create_sheet(cur[:31])
            for r in rows:
                ws.append(r)

    for line in lines:
        if line.strip().lower().startswith("### page"):
            flush()
            n = line.strip().split()[-1]
            cur, rows = f"page{n}", []
            continue
        cells = next(_csv.reader([line]))
        out_cells = []
        for c in cells:
            if keep_cell(c):
                out_cells.append(c)
            else:
                out_cells.append(None); dropped += 1
        rows.append(out_cells)
    flush()

    leftover = ([str(int(k) if float(k) == int(k) else k)
                 for k, c in budget["nums"].items() for _ in range(max(0, c))]
                + [k for k, c in budget["words"].items() for _ in range(max(0, c))]
                + add)
    if leftover:
        ws = wb.create_sheet("consensus_fix")
        for i in range(0, len(leftover), 8):
            ws.append(leftover[i:i + 8])
    if not wb.sheetnames:
        wb.create_sheet("page1")
    wb.save(form_dir / "golden.xlsx")
    gn, gw = xlsx_diff._atoms(xlsx_diff._cells(str(form_dir / "golden.xlsx")))
    print(f"{form_dir.name}: base={base_tag.split('__')[1][:24]} "
          f"convs={len(per_conv)} | golden {len(gn)}n/{len(gw)}w "
          f"(consensus {sum(cons['nums'].values())}n/"
          f"{sum(cons['words'].values())}w) | dropped {dropped} unsupported "
          f"cells, appended {len(leftover)}")


if __name__ == "__main__":
    cmd = sys.argv[1]
    base, apply, dirs, skip = DEFAULT_BASE, None, [], False
    for i, a in enumerate(sys.argv[2:], start=2):
        if skip:
            skip = False
            continue
        if a == "--base":
            base, skip = sys.argv[i + 1], True
        elif a == "--apply":
            apply, skip = sys.argv[i + 1], True
        elif not a.startswith("--"):
            dirs.append(Path(a))
    for d in dirs:
        if cmd == "report":
            report(d.resolve(), base)
        else:
            build(d.resolve(), base, apply)

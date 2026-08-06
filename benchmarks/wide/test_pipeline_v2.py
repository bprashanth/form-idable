#!/usr/bin/env python3
import tempfile
from pathlib import Path

import openpyxl
from openpyxl.styles import Border, Font, PatternFill, Side

import pipeline_v2


def main():
    assert pipeline_v2.reader_order("primary", "peer") == ["primary", "peer"]
    assert pipeline_v2.reader_order("primary", "") == ["primary"]
    assert pipeline_v2.reader_order("primary", "primary") == ["primary"]
    with tempfile.TemporaryDirectory() as temporary:
        root = Path(temporary)
        source, output = root / "source.xlsx", root / "output.xlsx"
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "page1"
        sheet["A1"] = "merged header"
        sheet["A1"].font = Font(bold=True, color="123456")
        sheet["A1"].fill = PatternFill("solid", fgColor="E7E6E6")
        thin = Side(style="thin", color="777777")
        sheet["A1"].border = Border(left=thin, right=thin, top=thin, bottom=thin)
        sheet.merge_cells("A1:B2")
        sheet.page_setup.orientation = "landscape"
        sheet.page_setup.fitToWidth = 1
        workbook.save(source)

        pipeline_v2.merge_workbooks([source], output)
        merged = openpyxl.load_workbook(output).active
        assert merged["A1"].value == "merged header"
        assert merged["A1"].font.bold
        assert merged["A1"].fill.fill_type == "solid"
        assert merged["A1"].border.left.style == "thin"
        assert {str(value) for value in merged.merged_cells.ranges} == {"A1:B2"}
        assert merged.page_setup.orientation == "landscape"
        assert merged.page_setup.fitToWidth == 1


if __name__ == "__main__":
    main()

#!/usr/bin/env python3
"""Independent artifact and quality audit for the production high sweep."""
from __future__ import annotations

import argparse
import json
import sys
from pathlib import Path

import fitz
import openpyxl
from PIL import Image

HERE = Path(__file__).resolve().parent
sys.path.insert(0, str(HERE))
import run_high_sweep  # noqa: E402
import wide_diff  # noqa: E402


def items(page):
    yield from page.get("metadata_fields") or []
    yield from page.get("free_text_regions") or []
    for table in page.get("tables") or []:
        for row in table.get("rows") or []:
            yield from row.get("cells") or []


def valid_bbox(box):
    return (isinstance(box, list) and len(box) == 4
            and all(isinstance(value, (int, float)) and 0 <= value <= 1 for value in box)
            and box[0] < box[2] and box[1] < box[3])


def norm(value):
    return "" if value is None else " ".join(str(value).strip().split())


def audit_artifacts(fixture: Path, output: Path) -> list[str]:
    errors = []
    document = json.loads((output / "canonical.json").read_text())
    review = json.loads((output / "review_manifest.json").read_text())
    analytics = json.loads((output / "analytics.json").read_text())
    crops = json.loads((output / "crops_manifest.json").read_text())
    pdf_pages = fitz.open(fixture / "input.pdf").page_count
    pages = document.get("pages") or []
    if len(pages) != pdf_pages:
        errors.append(f"canonical pages {len(pages)} != PDF pages {pdf_pages}")
    if analytics.get("summary", {}).get("pages") != pdf_pages:
        errors.append("analytics page count differs from PDF")

    workbook = openpyxl.load_workbook(output / "output.xlsx", data_only=True)
    expected_sheets = [f"page{number}" for number in range(1, pdf_pages + 1)] + ["ecology_review"]
    if workbook.sheetnames != expected_sheets:
        errors.append(f"workbook sheets {workbook.sheetnames!r} != {expected_sheets!r}")

    canonical_items = []
    for page in pages:
        page_number = page["page_number"]
        if not 1 <= page_number <= pdf_pages:
            errors.append(f"invalid canonical page {page_number}")
            continue
        sheet = workbook[f"page{page_number}"]
        for item in items(page):
            canonical_items.append(item)
            if not valid_bbox(item.get("bbox")):
                errors.append(f"page {page_number}: invalid bbox {item.get('bbox')}")
            row, column = item.get("xlsx_row"), item.get("xlsx_column")
            if not isinstance(row, int) or not isinstance(column, int):
                errors.append(f"page {page_number}: item lacks xlsx coordinate")
                continue
            cell = sheet.cell(row, column)
            if norm(cell.value) != norm(item.get("value")):
                errors.append(f"page {page_number} {row}:{column}: xlsx/canonical mismatch")

    if len(review.get("cells") or []) != len(canonical_items):
        errors.append("review manifest does not cover every canonical target")
    review_ids = [item.get("id") for item in review.get("cells") or []]
    if len(review_ids) != len(set(review_ids)):
        errors.append("duplicate review IDs")
    for item in [*(review.get("cells") or []),
                 *(review.get("views", {}).get("transcription_attention") or []),
                 *(review.get("views", {}).get("ecology_anomalies") or [])]:
        if item.get("bbox") is not None and not valid_bbox(item["bbox"]):
            errors.append(f"review item {item.get('id') or item.get('cell_id')}: invalid bbox")

    crop_pages = crops.get("pages") or []
    if len(crop_pages) != pdf_pages:
        errors.append("crop manifest page count differs from PDF")
    for page in crop_pages:
        render = output / "pages" / page["render"]
        if not render.exists():
            errors.append(f"missing render {render.name}")
        else:
            with Image.open(render) as image:
                if min(image.size) < 600:
                    errors.append(f"render {render.name} unexpectedly small: {image.size}")
        for crop in page.get("crops") or []:
            if not (output / "crops" / crop["file"]).exists():
                errors.append(f"missing crop {crop['file']}")
            if not valid_bbox(crop.get("bbox")):
                errors.append(f"crop {crop['file']}: invalid bbox")
    return errors


def extend_tokens(bucket, workbook):
    cells = wide_diff.xlsx_diff._cells(str(workbook))
    nums, words = wide_diff.xlsx_diff._atoms(cells)
    bucket["nums"].extend(nums)
    bucket["words"].extend(words)
    bucket["codes"].extend(wide_diff._semantic_codes(cells))


def aggregate(bucket):
    combined = ([str(value) for value in bucket["nums"]]
                + bucket["words"] + bucket["codes"])
    return bucket, combined


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--root", type=Path, default=run_high_sweep.DEFAULT_OUTPUT)
    parser.add_argument("--output", type=Path)
    args = parser.parse_args()
    fixtures = run_high_sweep.pdf_fixtures()
    token_sets = {name: {side: {kind: [] for kind in ("nums", "words", "codes")}
                         for side in ("golden", "candidate")}
                  for name in ("low", "high")}
    rows, errors = [], []
    for fixture in fixtures:
        output = args.root / fixture.name
        if not (output / "run.json").exists():
            errors.append(f"{fixture.name}: incomplete")
            continue
        content = output / "form/canonical_outputs/high_v1/output.xlsx"
        low = fixture / "codex_work/output.xlsx"
        high_metrics = run_high_sweep.compact(run_high_sweep.score(fixture, content))
        low_metrics = run_high_sweep.compact(run_high_sweep.score(fixture, low))
        artifact_errors = audit_artifacts(fixture, output)
        errors.extend(f"{fixture.name}: {message}" for message in artifact_errors)
        run = json.loads((output / "run.json").read_text())
        rows.append({"fixture": fixture.name,
                     "pages": fitz.open(fixture / "input.pdf").page_count,
                     "low": low_metrics, "high": high_metrics,
                     "semantic_f1_delta": round(high_metrics["semantic_all_f1"]
                                                - low_metrics["semantic_all_f1"], 3),
                     "review": run["review"],
                     "cost_usd": run["extraction"].get("cost_usd"),
                     "artifact_errors": artifact_errors})
        for name, candidate in (("low", low), ("high", content)):
            extend_tokens(token_sets[name]["golden"], fixture / "golden.xlsx")
            extend_tokens(token_sets[name]["candidate"], candidate)

    aggregates = {}
    for name, sides in token_sets.items():
        golden, golden_all = aggregate(sides["golden"])
        candidate, candidate_all = aggregate(sides["candidate"])
        aggregates[name] = {
            "semantic_all": wide_diff._prf(golden_all, candidate_all),
            "numeric": wide_diff._prf(golden["nums"], candidate["nums"]),
            "word": wide_diff._prf(golden["words"], candidate["words"]),
            "semantic_code": wide_diff._prf(golden["codes"], candidate["codes"]),
        }
    regressions = [row["fixture"] for row in rows if row["semantic_f1_delta"] < 0]
    targets = sum(row["review"].get("target_cells_including_blanks", 0) for row in rows)
    attention = sum(row["review"].get("transcription_review_cells", 0) for row in rows)
    ecology_flags = sum(row["review"].get("ecology_findings", 0) for row in rows)
    total_cost = round(sum(row["cost_usd"] or 0 for row in rows), 4)
    completed_pages = sum(row["pages"] for row in rows)
    report = {
        "version": "formidable-high-sweep-audit-v1",
        "expected_forms": len(fixtures), "completed_forms": len(rows),
        "expected_pages": sum(fitz.open(item / "input.pdf").page_count for item in fixtures),
        "completed_pages": completed_pages,
        "artifact_errors": errors, "per_form": rows, "aggregate": aggregates,
        "high_semantic_f1_regressions": regressions,
        "human_attention": {
            "targets_including_blanks": targets,
            "red_transcription_cells": attention,
            "red_fraction": round(attention / targets, 4) if targets else 0,
            "orange_ecology_cells": ecology_flags,
        },
        "total_model_cost_usd": total_cost,
        "model_cost_per_page_usd": round(total_cost / completed_pages, 4)
        if completed_pages else None,
    }
    destination = args.output or args.root / "audit.json"
    destination.write_text(json.dumps(report, indent=2) + "\n")
    print(json.dumps({key: report[key] for key in
                      ("expected_forms", "completed_forms", "expected_pages",
                       "completed_pages", "artifact_errors",
                       "high_semantic_f1_regressions", "human_attention",
                       "total_model_cost_usd", "model_cost_per_page_usd")}, indent=2))
    return 1 if errors else 0


if __name__ == "__main__":
    raise SystemExit(main())

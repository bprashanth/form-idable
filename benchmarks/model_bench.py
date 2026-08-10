#!/usr/bin/env python3
"""Model comparison experiment: find the cheapest model that matches codex CLI
on the TreePlots benchmark form.

Baseline = codex CLI (current prod impl): cell_frac~0.87, num_recall~0.72,
word_recall~0.46 against benchmarks/TreePlots20mx20m_merged.xlsx.

Scoring reuses the SAME tolerant diff the nightly regression uses (xlsx_diff.py),
so results are directly comparable to codex.

Providers:
  - gemini    : Google Generative Language API (direct key), thinking DISABLED
  - openrouter: OpenAI-compatible chat/completions with image inputs; real $ cost
                pulled from the /generation endpoint per call.

Modes:
  - oneshot : send all page overviews at once, ask for full transcription (cheapest)
  - crop    : agentic — model calls render_page (crop/zoom) tool, then transcribes
              (closer to codex behaviour; more tokens)

Usage:
  python3 run_experiment.py render                      # render page overviews once
  python3 run_experiment.py oneshot <provider> <model>  # one model, one-shot
  python3 run_experiment.py crop    <provider> <model>  # one model, agentic crops
"""
import base64, csv, io, json, os, sys, time, urllib.request, urllib.error
from pathlib import Path

HERE   = Path(__file__).parent
FORMID = Path.home() / "src/github.com/bprashanth/form-idable"
GSHEP  = Path.home() / "src/github.com/bprashanth/good-shepherd/agents/formidable"
PDF    = FORMID / "benchmarks/TreePlots20mx20m.pdf"
GOLDEN = FORMID / "benchmarks/TreePlots20mx20m_merged.xlsx"
RENDER = GSHEP / "tools/render_page.py"
PAGES_DIR = HERE / "pages"
OUT_DIR   = HERE / "outputs"
CFG    = Path.home() / ".config/formidable"

sys.path.insert(0, str(GSHEP))          # for xlsx_diff
import xlsx_diff                          # noqa: E402
import openpyxl                           # noqa: E402

# ── keys ──────────────────────────────────────────────────────────
def _key(name):
    return json.loads((CFG / f"{name}.json").read_text())["api_key"]

# ── the transcription prompt (adapted from prompts/codex_prompt.md) ──
TRANSCRIBE_PROMPT = """You are transcribing scanned handwritten ecological field
datasheets (Western Ghats 20x20m tree-plot surveys) into tabular text.

You are given page overview images of a multi-page form. Transcribe EVERYTHING
visible on every page: metadata headers (date, site/plot ID, GPS/coordinates,
collectors), every table (species, GBH, height, counts), tally/tick marks, and
marginal notes.

Notation rules:
- a dot/period in a cell means the value 0
- a continuous line through a cell means "no entry" (leave blank)
- tally marks (I, l, |) are a count — sum to an integer
- ticks/X/checkmarks mean present -> transcribe as X

Output ONLY the transcription as CSV — one table row per line, cells separated by
commas. Put a line "### PAGE N" before each page's content. Do not add prose,
explanations, or markdown fences. Transcribe values as literally as you can read
them; it is better to include an uncertain value than to omit it.
"""

# ── rendering ─────────────────────────────────────────────────────
def render_pages():
    import subprocess
    PAGES_DIR.mkdir(parents=True, exist_ok=True)
    import fitz
    n = fitz.open(str(PDF)).page_count
    out = []
    for p in range(1, n + 1):
        dst = PAGES_DIR / f"page_{p}.png"
        subprocess.run([sys.executable, str(RENDER), str(PDF),
                        "--out", str(dst), "--page", str(p), "--zoom", "3"],
                       check=True, capture_output=True)
        out.append(dst)
    print(f"rendered {len(out)} pages -> {PAGES_DIR}")
    return out

def _page_pngs():
    return sorted(PAGES_DIR.glob("page_*.png"))

def render_tiles():
    """Split each page into top+bottom halves, each rendered to the 1568px cap —
    ~2x the vertical detail of a full-page overview. The deterministic stand-in
    for codex's crop/zoom, tested as an enriched one-shot."""
    import subprocess, fitz
    TILES_DIR = HERE / "tiles"; TILES_DIR.mkdir(parents=True, exist_ok=True)
    n = fitz.open(str(PDF)).page_count
    out = []
    for p in range(1, n + 1):
        for half, (y0, y1) in enumerate([(0.0, 0.55), (0.45, 1.0)]):  # slight overlap
            dst = TILES_DIR / f"page_{p}_h{half}.png"
            subprocess.run([sys.executable, str(RENDER), str(PDF), "--out", str(dst),
                            "--page", str(p), "--bbox", f"0,{y0},1,{y1}", "--zoom", "6"],
                           check=True, capture_output=True)
            out.append(dst)
    print(f"rendered {len(out)} tiles -> {TILES_DIR}")
    return out

def _tile_pngs():
    return sorted((HERE / "tiles").glob("page_*_h*.png"))

def _b64(path):
    return base64.b64encode(Path(path).read_bytes()).decode()

# ── output parsing -> xlsx ────────────────────────────────────────
def text_to_xlsx(text, dst):
    wb = openpyxl.Workbook(); wb.remove(wb.active)
    page, rows = "v2", []
    def flush(name, rws):
        if not rws: return
        ws = wb.create_sheet(title=name[:31])
        for r in rws:
            ws.append(r)
    cur = "v2"
    for line in text.splitlines():
        if line.strip().lower().startswith("### page"):
            flush(cur, rows); rows = []; cur = line.strip().split("###")[-1].strip() or "page"
            continue
        rows.append(next(csv.reader([line])))
    flush(cur, rows)
    if not wb.sheetnames:
        wb.create_sheet("v2")
    wb.save(dst)
    return dst

# ── HTTP helper ───────────────────────────────────────────────────
def _post(url, payload, headers, timeout=300):
    req = urllib.request.Request(url, data=json.dumps(payload).encode(),
                                 headers={"Content-Type": "application/json", **headers})
    with urllib.request.urlopen(req, timeout=timeout) as r:
        return json.loads(r.read().decode())

# ── providers: one-shot ───────────────────────────────────────────
def gemini_oneshot(model, pages):
    key = _key("gemini")
    parts = [{"text": TRANSCRIBE_PROMPT}]
    for p in pages:
        parts.append({"inline_data": {"mime_type": "image/png", "data": _b64(p)}})
    payload = {
        "contents": [{"role": "user", "parts": parts}],
        "generationConfig": {"temperature": 0, "thinkingConfig": {"thinkingBudget": 0}},
    }
    url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={key}"
    t0 = time.time()
    resp = _post(url, payload, {})
    dt = time.time() - t0
    text = "".join(part.get("text", "")
                   for part in resp["candidates"][0]["content"]["parts"])
    um = resp.get("usageMetadata", {})
    cost = _gemini_cost(model, um)
    return text, {"in_tok": um.get("promptTokenCount"), "out_tok": um.get("candidatesTokenCount"),
                  "cost_usd": cost, "latency_s": round(dt, 1)}

def openrouter_oneshot(model, pages):
    key = _key("openrouter")
    content = [{"type": "text", "text": TRANSCRIBE_PROMPT}]
    for p in pages:
        content.append({"type": "image_url",
                        "image_url": {"url": f"data:image/png;base64,{_b64(p)}"}})
    payload = {"model": model, "temperature": 0,
               "messages": [{"role": "user", "content": content}],
               "reasoning": {"enabled": False}, "usage": {"include": True}}
    hdrs = {"Authorization": f"Bearer {key}",
            "HTTP-Referer": "https://formidable.local", "X-Title": "formidable-eval"}
    t0 = time.time()
    try:
        resp = _post("https://openrouter.ai/api/v1/chat/completions", payload, hdrs)
    except urllib.error.HTTPError as e:
        # Some models (e.g. gpt-5-nano) require reasoning — retry without disabling it.
        if e.code == 400 and b"easoning" in e.read():
            payload.pop("reasoning", None)
            resp = _post("https://openrouter.ai/api/v1/chat/completions", payload, hdrs)
        else:
            raise
    dt = time.time() - t0
    text = resp["choices"][0]["message"]["content"]
    usage = resp.get("usage", {})
    cost = usage.get("cost")
    if cost is None:
        cost = _openrouter_cost(resp.get("id"), key)
    return text, {"in_tok": usage.get("prompt_tokens"), "out_tok": usage.get("completion_tokens"),
                  "cost_usd": cost, "latency_s": round(dt, 1)}

# ── cost helpers ──────────────────────────────────────────────────
# Gemini flash pricing (per 1M tokens) — approximate, update if models change.
_GEMINI_PRICES = {
    "gemini-2.5-flash":       (0.30, 2.50),
    "gemini-2.0-flash":       (0.10, 0.40),
    "gemini-flash-latest":    (0.30, 2.50),
}
def _gemini_cost(model, um):
    pin, pout = _GEMINI_PRICES.get(model, (0.30, 2.50))
    it = (um.get("promptTokenCount") or 0) / 1e6
    ot = (um.get("candidatesTokenCount") or 0) / 1e6
    return round(it * pin + ot * pout, 5)

def _openrouter_cost(gen_id, key):
    if not gen_id: return None
    try:
        time.sleep(1)
        req = urllib.request.Request(
            f"https://openrouter.ai/api/v1/generation?id={gen_id}",
            headers={"Authorization": f"Bearer {key}"})
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read().decode())["data"].get("total_cost")
    except Exception:
        return None

# ── driver ────────────────────────────────────────────────────────
def run_one(provider, model, mode="oneshot"):
    images = _tile_pngs() if mode == "tiled" else _page_pngs()
    assert images, "run `render`/`tiles` first"
    sender = {"gemini": gemini_oneshot, "openrouter": openrouter_oneshot}.get(provider)
    if not sender:
        raise SystemExit(f"unsupported provider {provider}")
    fn = lambda model, _pages: sender(model, images)  # noqa: E731
    OUT_DIR.mkdir(parents=True, exist_ok=True)
    tag = f"{provider}__{model.replace('/', '_')}__{mode}"
    try:
        text, meta = fn(model, images)
    except urllib.error.HTTPError as e:
        print(json.dumps({"model": model, "provider": provider, "mode": mode,
                          "error": f"HTTP {e.code}: {e.read().decode()[:300]}"}))
        return
    (OUT_DIR / f"{tag}.txt").write_text(text)
    xlsx = text_to_xlsx(text, OUT_DIR / f"{tag}.xlsx")
    result = xlsx_diff.compare(str(GOLDEN), str(xlsx))
    m = result["metrics"]
    row = {"model": model, "provider": provider, "mode": mode,
           "passed": result["passed"],
           "cell_frac": m["cell_frac"], "num_recall": m["num_recall"],
           "word_recall": m["word_recall"], **meta}
    (OUT_DIR / f"{tag}.json").write_text(json.dumps(row, indent=2))
    print(json.dumps(row))
    return row

if __name__ == "__main__":
    cmd = sys.argv[1] if len(sys.argv) > 1 else "render"
    if cmd == "render":
        render_pages()
    elif cmd == "tiles":
        render_tiles()
    else:
        run_one(sys.argv[2], sys.argv[3], cmd)
